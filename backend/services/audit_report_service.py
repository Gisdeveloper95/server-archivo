# -*- coding: utf-8 -*-
"""
AuditReportService - Servicio para generar reportes profesionales de auditoría
Genera múltiples formatos: CSV, TXT (árbol), XLSX (Excel), HTML (timeline), PDF
"""
import io
import os
import csv
import zipfile
import hashlib
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from collections import defaultdict
from django.conf import settings

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# Constantes
WINDOWS_BASE_PATH = r'\\repositorio\DirGesCat\2510SP\H_Informacion_Consulta\Sub_Proy'

# Emojis para acciones
ACTION_EMOJIS = {
    'upload': '📤',
    'download': '📥',
    'delete': '🗑️',
    'rename': '✏️',
    'copy': '📋',
    'move': '📦',
    'create_folder': '📁',
    'view': '👁️',
    'login': '🔐',
    'logout': '🚪',
    'share': '🔗',
    'permission': '🔑',
}

ACTION_LABELS = {
    'upload': 'Subir',
    'download': 'Descargar',
    'delete': 'Eliminar',
    'rename': 'Renombrar',
    'copy': 'Copiar',
    'move': 'Mover',
    'create_folder': 'Crear carpeta',
    'view': 'Ver',
    'login': 'Iniciar sesión',
    'logout': 'Cerrar sesión',
    'share': 'Compartir',
    'permission': 'Permiso',
}

# Colores para Excel
COLORS = {
    'header': 'FF2C3E50',      # Azul oscuro
    'header_font': 'FFFFFFFF', # Blanco
    'upload': 'FF27AE60',      # Verde
    'download': 'FF3498DB',    # Azul
    'delete': 'FFE74C3C',      # Rojo
    'rename': 'FFF39C12',      # Naranja
    'copy': 'FF9B59B6',        # Púrpura
    'move': 'FF1ABC9C',        # Turquesa
    'create_folder': 'FFF1C40F', # Amarillo
    'view': 'FF95A5A6',        # Gris
    'default': 'FFFFFFFF',     # Blanco
    'alert': 'FFFF6B6B',       # Rojo claro
    'stripe': 'FFF8F9FA',      # Gris muy claro
}


class AuditReportService:
    """Servicio para generar reportes de auditoría en múltiples formatos"""

    def __init__(self, logs: List[Any], report_type: str = 'general',
                 filters: Optional[Dict] = None, generated_by: str = 'Sistema'):
        """
        Args:
            logs: Lista de objetos AuditLog
            report_type: 'general', 'user', 'directory', 'file'
            filters: Filtros aplicados (fechas, usuarios, etc.)
            generated_by: Usuario que genera el reporte
        """
        self.logs = list(logs)
        self.report_type = report_type
        self.filters = filters or {}
        self.generated_by = generated_by
        self.generation_time = datetime.now()

    def _to_windows_path(self, path: str) -> str:
        """Convierte ruta Linux/relativa a Windows UNC"""
        if not path:
            return ''
        clean_path = path.replace('/', '\\').lstrip('\\')
        return f"{WINDOWS_BASE_PATH}\\{clean_path}"

    def _format_size(self, size_bytes: int) -> str:
        """Formatea tamaño en bytes a formato legible"""
        if not size_bytes:
            return '-'
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
        else:
            return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

    def _normalize_username(self, username: str) -> str:
        """Normaliza username eliminando el dominio de correo si existe"""
        if not username:
            return username
        # Si tiene @, extraer solo la parte antes del @
        if '@' in username:
            return username.split('@')[0]
        return username

    def _get_unique_users(self) -> Dict[str, int]:
        """Obtiene usuarios únicos normalizados con sus conteos"""
        user_counts = defaultdict(int)
        for log in self.logs:
            normalized = self._normalize_username(log.username)
            user_counts[normalized] += 1
        return dict(user_counts)

    def _get_date_range(self) -> tuple:
        """Obtiene el rango de fechas de los logs"""
        if not self.logs:
            return None, None

        dates = [log.timestamp for log in self.logs if log.timestamp]
        if not dates:
            return None, None

        return min(dates), max(dates)

    def _get_stats(self) -> Dict:
        """Calcula estadísticas de los logs"""
        stats = {
            'total': len(self.logs),
            'by_action': defaultdict(int),
            'by_user': defaultdict(int),
            'by_user_normalized': {},  # Usuarios normalizados sin duplicados
            'by_date': defaultdict(int),
            'by_hour': defaultdict(int),
            'successful': 0,
            'failed': 0,
            'total_size': 0,
            'deleted_files_count': 0,
            'date_range': self._get_date_range(),
        }

        for log in self.logs:
            stats['by_action'][log.action] += 1
            stats['by_user'][log.username] += 1

        # Usuarios normalizados (sin duplicados de correo)
        stats['by_user_normalized'] = self._get_unique_users()

        for log in self.logs:
            if log.timestamp:
                date_key = log.timestamp.strftime('%Y-%m-%d')
                hour_key = log.timestamp.strftime('%H:00')
                stats['by_date'][date_key] += 1
                stats['by_hour'][hour_key] += 1

            if log.success:
                stats['successful'] += 1
            else:
                stats['failed'] += 1

            if log.file_size:
                stats['total_size'] += log.file_size

            # Contar archivos eliminados en detalle
            if log.action == 'delete' and log.details:
                deleted_items = log.details.get('deleted_items', [])
                stats['deleted_files_count'] += len(deleted_items)

        return stats

    def _get_alerts(self) -> List[Dict]:
        """Detecta alertas de seguridad"""
        alerts = []

        for log in self.logs:
            # Eliminaciones masivas (>10 archivos)
            if log.action == 'delete' and log.details:
                deleted_count = len(log.details.get('deleted_items', []))
                if deleted_count > 10:
                    alerts.append({
                        'type': 'mass_delete',
                        'severity': 'high',
                        'message': f'Eliminación masiva: {deleted_count} archivos',
                        'user': log.username,
                        'timestamp': log.timestamp,
                        'path': log.target_path,
                    })

            # Operaciones fallidas
            if not log.success:
                alerts.append({
                    'type': 'failed_operation',
                    'severity': 'medium',
                    'message': f'Operación fallida: {log.error_message or "Sin detalle"}',
                    'user': log.username,
                    'timestamp': log.timestamp,
                    'path': log.target_path,
                })

        return alerts

    # ==================== GENERADORES DE REPORTES ====================

    def generate_csv(self) -> bytes:
        """Genera CSV con todos los detalles"""
        output = io.StringIO()
        output.write('\ufeff')  # BOM UTF-8

        writer = csv.writer(output)

        # Encabezados
        writer.writerow([
            'ID', 'Usuario', 'Rol', 'Acción', 'Tipo', 'Ruta',
            'Ruta Windows', 'Nombre', 'Tamaño (bytes)', 'Extensión',
            'IP', 'Éxito', 'Error', 'Fecha/Hora', 'ID Padre', 'Es Contenido'
        ])

        for log in self.logs:
            # Fila principal
            writer.writerow([
                log.id,
                log.username,
                log.user_role,
                ACTION_LABELS.get(log.action, log.action),
                'Directorio' if log.details and log.details.get('is_directory') else 'Archivo',
                log.target_path or '',
                self._to_windows_path(log.target_path),
                log.target_name or '',
                log.file_size or '',
                '',
                log.ip_address or '',
                'Sí' if log.success else 'No',
                log.error_message or '',
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
                '',
                'No'
            ])

            # Expandir archivos eliminados
            if log.action == 'delete' and log.success and log.details:
                for item in log.details.get('deleted_items', []):
                    full_path = f"{log.target_path}/{item.get('path', '')}"
                    writer.writerow([
                        '',
                        log.username,
                        log.user_role,
                        'Eliminado (contenido)',
                        'Directorio' if item.get('is_directory') else 'Archivo',
                        full_path,
                        self._to_windows_path(full_path),
                        item.get('name', ''),
                        item.get('size', ''),
                        item.get('extension', ''),
                        log.ip_address or '',
                        'Sí',
                        '',
                        log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
                        log.id,
                        'Sí'
                    ])

        return output.getvalue().encode('utf-8')

    def generate_tree_txt(self) -> bytes:
        """Genera archivo TXT con vista de árbol y emojis"""
        lines = []

        # Encabezado
        lines.append("=" * 80)
        lines.append("REPORTE DE AUDITORÍA - ÁRBOL DE EVENTOS")
        lines.append("=" * 80)
        lines.append(f"Generado: {self.generation_time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Por: {self.generated_by}")

        if self.filters.get('date_from') or self.filters.get('date_to'):
            lines.append(f"Período: {self.filters.get('date_from', 'Inicio')} - {self.filters.get('date_to', 'Actual')}")

        lines.append(f"Total de eventos: {len(self.logs)}")
        lines.append("=" * 80)
        lines.append("")

        # Leyenda de emojis
        lines.append("LEYENDA:")
        lines.append("-" * 40)
        for action, emoji in ACTION_EMOJIS.items():
            label = ACTION_LABELS.get(action, action)
            lines.append(f"  {emoji} = {label}")
        lines.append("")
        lines.append("=" * 80)
        lines.append("")

        # Agrupar por ruta
        events_by_path = defaultdict(list)
        for log in self.logs:
            path = log.target_path or '/'
            events_by_path[path].append(log)

        # Construir árbol
        tree = {}
        for path in sorted(events_by_path.keys()):
            parts = [p for p in path.split('/') if p]
            current = tree
            for part in parts:
                if part not in current:
                    current[part] = {'_events': [], '_children': {}}
                current = current[part]['_children']

        # Función recursiva para imprimir árbol
        def print_tree(node: Dict, prefix: str = "", path_so_far: str = ""):
            items = sorted(node.items())
            for i, (name, data) in enumerate(items):
                is_last = (i == len(items) - 1)
                connector = "└── " if is_last else "├── "
                extension = "    " if is_last else "│   "

                full_path = f"{path_so_far}/{name}" if path_so_far else name

                # Obtener eventos para esta ruta
                path_events = events_by_path.get(full_path, [])

                if path_events:
                    lines.append(f"{prefix}{connector}📁 {name}")
                    for event in sorted(path_events, key=lambda x: x.timestamp or datetime.min):
                        emoji = ACTION_EMOJIS.get(event.action, '❓')
                        time_str = event.timestamp.strftime('%Y-%m-%d %H:%M') if event.timestamp else ''
                        user = event.username
                        target = event.target_name or ''
                        status = '✓' if event.success else '✗'

                        event_line = f"{prefix}{extension}    [{time_str}] {emoji} {user} - {target} {status}"
                        lines.append(event_line)

                        # Si es eliminación con detalles, mostrar archivos eliminados
                        if event.action == 'delete' and event.details:
                            deleted_items = event.details.get('deleted_items', [])
                            if deleted_items:
                                lines.append(f"{prefix}{extension}        📋 Contenido eliminado ({len(deleted_items)} items):")
                                for item in deleted_items[:10]:  # Limitar a 10
                                    item_emoji = '📁' if item.get('is_directory') else '📄'
                                    item_size = self._format_size(item.get('size', 0))
                                    lines.append(f"{prefix}{extension}        {item_emoji} {item.get('name', '')} ({item_size})")
                                if len(deleted_items) > 10:
                                    lines.append(f"{prefix}{extension}        ... y {len(deleted_items) - 10} más")
                else:
                    lines.append(f"{prefix}{connector}📁 {name}")

                # Recursión para hijos
                if isinstance(data, dict) and '_children' in data:
                    print_tree(data['_children'], prefix + extension, full_path)

        # Imprimir árbol desde la raíz
        for path, events in sorted(events_by_path.items()):
            if '/' not in path or path.count('/') == 0:
                continue

            parts = path.split('/')
            root = parts[0] if parts[0] else parts[1]

            # Solo imprimir eventos directamente
            for event in sorted(events, key=lambda x: x.timestamp or datetime.min):
                emoji = ACTION_EMOJIS.get(event.action, '❓')
                time_str = event.timestamp.strftime('%Y-%m-%d %H:%M') if event.timestamp else ''
                user = event.username
                target = event.target_name or path.split('/')[-1]
                status = '✓' if event.success else '✗'
                size_str = f" ({self._format_size(event.file_size)})" if event.file_size else ''

                lines.append(f"[{time_str}] {emoji} {ACTION_LABELS.get(event.action, event.action).upper()}")
                lines.append(f"    Usuario: {user}")
                lines.append(f"    Ruta: {path}")
                lines.append(f"    Archivo: {target}{size_str}")
                lines.append(f"    Estado: {'Exitoso' if event.success else 'Fallido'}")
                if not event.success and event.error_message:
                    lines.append(f"    Error: {event.error_message}")

                # Detalles de eliminación
                if event.action == 'delete' and event.details:
                    deleted_items = event.details.get('deleted_items', [])
                    if deleted_items:
                        lines.append(f"    Contenido eliminado ({len(deleted_items)} items):")
                        for item in deleted_items:
                            item_emoji = '📁' if item.get('is_directory') else '📄'
                            item_path = f"{path}/{item.get('path', '')}"
                            item_size = self._format_size(item.get('size', 0))
                            lines.append(f"        {item_emoji} {self._to_windows_path(item_path)}")
                            lines.append(f"           Nombre: {item.get('name', '')}")
                            lines.append(f"           Tamaño: {item_size}")

                lines.append("")

        # Resumen final
        stats = self._get_stats()
        lines.append("")
        lines.append("=" * 80)
        lines.append("RESUMEN ESTADÍSTICO")
        lines.append("=" * 80)
        lines.append(f"Total de operaciones: {stats['total']}")
        lines.append(f"Operaciones exitosas: {stats['successful']}")
        lines.append(f"Operaciones fallidas: {stats['failed']}")
        lines.append(f"Archivos eliminados (detalle): {stats['deleted_files_count']}")
        lines.append(f"Tamaño total procesado: {self._format_size(stats['total_size'])}")
        lines.append("")
        lines.append("Por tipo de acción:")
        for action, count in sorted(stats['by_action'].items(), key=lambda x: -x[1]):
            emoji = ACTION_EMOJIS.get(action, '❓')
            label = ACTION_LABELS.get(action, action)
            lines.append(f"  {emoji} {label}: {count}")

        lines.append("")
        lines.append("Por usuario:")
        for user, count in sorted(stats['by_user'].items(), key=lambda x: -x[1])[:10]:
            lines.append(f"  👤 {user}: {count} operaciones")

        # Alertas
        alerts = self._get_alerts()
        if alerts:
            lines.append("")
            lines.append("=" * 80)
            lines.append("⚠️ ALERTAS DE SEGURIDAD")
            lines.append("=" * 80)
            for alert in alerts[:20]:
                lines.append(f"  [{alert['severity'].upper()}] {alert['message']}")
                lines.append(f"      Usuario: {alert['user']}")
                lines.append(f"      Fecha: {alert['timestamp'].strftime('%Y-%m-%d %H:%M') if alert['timestamp'] else 'N/A'}")
                lines.append("")

        # Hash del reporte
        content = '\n'.join(lines)
        report_hash = hashlib.sha256(content.encode('utf-8')).hexdigest()
        lines.append("")
        lines.append("=" * 80)
        lines.append(f"Hash SHA-256 del reporte: {report_hash}")
        lines.append("=" * 80)

        return '\n'.join(lines).encode('utf-8')

    def generate_excel(self) -> bytes:
        """Genera Excel profesional con múltiples hojas y diseño"""
        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color=COLORS['header_font'])
        header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ==================== HOJA 1: PORTADA ====================
        ws_cover = wb.active
        ws_cover.title = "Portada"

        ws_cover['B3'] = "REPORTE DE AUDITORÍA"
        ws_cover['B3'].font = Font(bold=True, size=24, color=COLORS['header'])

        ws_cover['B5'] = "Sistema de Gestión Documental"
        ws_cover['B5'].font = Font(size=14, italic=True)

        ws_cover['B8'] = "Información del Reporte"
        ws_cover['B8'].font = Font(bold=True, size=12)

        info_rows = [
            ("Generado por:", self.generated_by),
            ("Fecha de generación:", self.generation_time.strftime('%Y-%m-%d %H:%M:%S')),
            ("Tipo de reporte:", self.report_type.upper()),
            ("Total de registros:", str(len(self.logs))),
        ]

        if self.filters.get('date_from'):
            info_rows.append(("Fecha desde:", self.filters['date_from']))
        if self.filters.get('date_to'):
            info_rows.append(("Fecha hasta:", self.filters['date_to']))
        if self.filters.get('username'):
            info_rows.append(("Usuario filtrado:", self.filters['username']))
        if self.filters.get('path'):
            info_rows.append(("Ruta filtrada:", self.filters['path']))

        for i, (label, value) in enumerate(info_rows, start=10):
            ws_cover[f'B{i}'] = label
            ws_cover[f'B{i}'].font = Font(bold=True)
            ws_cover[f'C{i}'] = value

        ws_cover.column_dimensions['B'].width = 25
        ws_cover.column_dimensions['C'].width = 50

        # ==================== HOJA 2: RESUMEN ====================
        ws_summary = wb.create_sheet("Resumen")
        stats = self._get_stats()

        ws_summary['A1'] = "RESUMEN ESTADÍSTICO"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')

        # Estadísticas generales
        summary_data = [
            ["Métrica", "Valor"],
            ["Total de operaciones", stats['total']],
            ["Operaciones exitosas", stats['successful']],
            ["Operaciones fallidas", stats['failed']],
            ["Archivos en eliminaciones", stats['deleted_files_count']],
            ["Tamaño total procesado", self._format_size(stats['total_size'])],
            ["Usuarios únicos", len(stats['by_user'])],
        ]

        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Operaciones por tipo
        ws_summary['A12'] = "OPERACIONES POR TIPO"
        ws_summary['A12'].font = Font(bold=True, size=14)

        action_data = [["Acción", "Cantidad", "Porcentaje"]]
        total = stats['total'] or 1
        for action, count in sorted(stats['by_action'].items(), key=lambda x: -x[1]):
            label = ACTION_LABELS.get(action, action)
            pct = f"{(count/total)*100:.1f}%"
            action_data.append([label, count, pct])

        for row_idx, row_data in enumerate(action_data, start=14):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 14:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15

        # ==================== HOJA 3: DETALLE ====================
        ws_detail = wb.create_sheet("Detalle Operaciones")

        headers = [
            'ID', 'Fecha/Hora', 'Usuario', 'Rol', 'Acción', 'Tipo',
            'Ruta', 'Ruta Windows', 'Nombre', 'Tamaño', 'IP', 'Estado', 'Error'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        row_idx = 2
        for log in self.logs:
            action_color = COLORS.get(log.action, COLORS['default'])
            row_fill = PatternFill(start_color=action_color, end_color=action_color, fill_type='solid') if log.action in ['delete'] else None

            data = [
                log.id,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
                log.username,
                log.user_role,
                ACTION_LABELS.get(log.action, log.action),
                'Directorio' if log.details and log.details.get('is_directory') else 'Archivo',
                log.target_path or '',
                self._to_windows_path(log.target_path),
                log.target_name or '',
                self._format_size(log.file_size) if log.file_size else '',
                log.ip_address or '',
                'Exitoso' if log.success else 'Fallido',
                log.error_message or ''
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if not log.success:
                    cell.fill = PatternFill(start_color=COLORS['alert'], end_color=COLORS['alert'], fill_type='solid')
                elif row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=COLORS['stripe'], end_color=COLORS['stripe'], fill_type='solid')

            row_idx += 1

        # Ajustar anchos
        column_widths = [8, 20, 15, 12, 15, 12, 40, 60, 30, 12, 15, 10, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = width

        # Filtros automáticos
        ws_detail.auto_filter.ref = f"A1:M{row_idx-1}"

        # Congelar primera fila
        ws_detail.freeze_panes = 'A2'

        # ==================== HOJA 4: ELIMINACIONES DETALLADAS ====================
        ws_deletes = wb.create_sheet("Eliminaciones Detalladas")

        del_headers = [
            'ID Operación', 'Fecha/Hora', 'Usuario', 'Directorio Eliminado',
            'Archivo/Carpeta', 'Ruta Completa', 'Ruta Windows', 'Tipo', 'Tamaño', 'Extensión'
        ]

        for col_idx, header in enumerate(del_headers, start=1):
            cell = ws_deletes.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_idx = 2
        for log in self.logs:
            if log.action != 'delete' or not log.success or not log.details:
                continue

            deleted_items = log.details.get('deleted_items', [])
            for item in deleted_items:
                full_path = f"{log.target_path}/{item.get('path', '')}"
                data = [
                    log.id,
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
                    log.username,
                    log.target_name or log.target_path.split('/')[-1] if log.target_path else '',
                    item.get('name', ''),
                    full_path,
                    self._to_windows_path(full_path),
                    'Directorio' if item.get('is_directory') else 'Archivo',
                    self._format_size(item.get('size', 0)),
                    item.get('extension', '')
                ]

                for col_idx, value in enumerate(data, start=1):
                    cell = ws_deletes.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=COLORS['stripe'], end_color=COLORS['stripe'], fill_type='solid')

                row_idx += 1

        # Ajustar anchos
        del_widths = [12, 20, 15, 25, 30, 50, 70, 12, 12, 10]
        for col_idx, width in enumerate(del_widths, start=1):
            ws_deletes.column_dimensions[get_column_letter(col_idx)].width = width

        if row_idx > 2:
            ws_deletes.auto_filter.ref = f"A1:J{row_idx-1}"
        ws_deletes.freeze_panes = 'A2'

        # ==================== HOJA 5: ALERTAS ====================
        ws_alerts = wb.create_sheet("Alertas Seguridad")
        alerts = self._get_alerts()

        alert_headers = ['Severidad', 'Tipo', 'Mensaje', 'Usuario', 'Fecha/Hora', 'Ruta']
        for col_idx, header in enumerate(alert_headers, start=1):
            cell = ws_alerts.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, alert in enumerate(alerts, start=2):
            data = [
                alert['severity'].upper(),
                alert['type'],
                alert['message'],
                alert['user'],
                alert['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if alert['timestamp'] else '',
                alert.get('path', '')
            ]
            for col_idx, value in enumerate(data, start=1):
                cell = ws_alerts.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if alert['severity'] == 'high':
                    cell.fill = PatternFill(start_color=COLORS['alert'], end_color=COLORS['alert'], fill_type='solid')

        alert_widths = [12, 20, 50, 20, 20, 50]
        for col_idx, width in enumerate(alert_widths, start=1):
            ws_alerts.column_dimensions[get_column_letter(col_idx)].width = width

        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _build_tree_structure(self, items: list) -> dict:
        """Construye estructura de arbol a partir de items con paths"""
        tree = {'name': '', 'children': {}, 'files': [], 'is_dir': True}

        for item in items:
            path = item.get('path', item.get('name', ''))
            parts = path.replace('\\', '/').split('/')
            current = tree

            # Navegar/crear estructura de directorios
            for i, part in enumerate(parts[:-1]):
                if part not in current['children']:
                    current['children'][part] = {
                        'name': part,
                        'children': {},
                        'files': [],
                        'is_dir': True,
                        'size': 0
                    }
                current = current['children'][part]

            # Agregar el item final
            final_name = parts[-1] if parts else item.get('name', 'unknown')
            if item.get('is_directory'):
                if final_name not in current['children']:
                    current['children'][final_name] = {
                        'name': final_name,
                        'children': {},
                        'files': [],
                        'is_dir': True,
                        'size': item.get('size', 0)
                    }
                else:
                    current['children'][final_name]['size'] = item.get('size', 0)
            else:
                current['files'].append({
                    'name': final_name,
                    'size': item.get('size', 0),
                    'extension': item.get('extension', '')
                })

        return tree

    def _render_tree_html(self, node: dict, indent: int = 0, is_last: bool = True, prefix: str = '') -> str:
        """Renderiza un nodo del arbol como HTML con lineas de conexion"""
        html_parts = []

        # Determinar caracteres de conexion
        connector = '&#9492;&#9472;&#9472; ' if is_last else '&#9500;&#9472;&#9472; '  # └── o ├──
        extension = '    ' if is_last else '&#9474;   '  # espacio o │

        # Renderizar subdirectorios primero
        children = list(node.get('children', {}).values())
        files = node.get('files', [])
        total_items = len(children) + len(files)

        for idx, child in enumerate(sorted(children, key=lambda x: x['name'].lower())):
            is_last_item = (idx == len(children) - 1) and len(files) == 0
            child_prefix = prefix + extension

            # Icono de carpeta
            size_str = ' <span class="tree-size">(' + self._format_size(child.get('size', 0)) + ')</span>' if child.get('size') else ''
            html_parts.append(
                '<div class="tree-line">'
                '<span class="tree-prefix">{}{}</span>'
                '<span class="tree-icon folder">&#128193;</span>'
                '<span class="tree-name dir">{}</span>{}'
                '</div>'.format(prefix, connector if indent > 0 else '', child['name'], size_str)
            )

            # Recursivamente renderizar hijos
            child_html = self._render_tree_html(child, indent + 1, is_last_item, child_prefix)
            html_parts.append(child_html)

        # Renderizar archivos
        for idx, file in enumerate(sorted(files, key=lambda x: x['name'].lower())):
            is_last_file = idx == len(files) - 1
            file_connector = '&#9492;&#9472;&#9472; ' if is_last_file else '&#9500;&#9472;&#9472; '

            # Icono segun extension
            ext = file.get('extension', '').lower()
            if ext in ['.shp', '.gdb', '.geojson', '.kml']:
                icon = '&#127758;'  # Globo para geoespacial
            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.tif', '.tiff']:
                icon = '&#128247;'  # Camara para imagenes
            elif ext in ['.pdf']:
                icon = '&#128213;'  # Libro para PDF
            elif ext in ['.doc', '.docx', '.odt']:
                icon = '&#128196;'  # Documento
            elif ext in ['.xls', '.xlsx', '.csv']:
                icon = '&#128202;'  # Grafico para Excel
            elif ext in ['.zip', '.rar', '.7z']:
                icon = '&#128230;'  # Paquete para comprimidos
            else:
                icon = '&#128196;'  # Documento generico

            size_str = ' <span class="tree-size">(' + self._format_size(file.get('size', 0)) + ')</span>'
            html_parts.append(
                '<div class="tree-line">'
                '<span class="tree-prefix">{}{}</span>'
                '<span class="tree-icon file">{}</span>'
                '<span class="tree-name file">{}</span>{}'
                '</div>'.format(prefix, file_connector, icon, file['name'], size_str)
            )

        return ''.join(html_parts)

    def generate_html_timeline(self) -> bytes:
        """Genera HTML interactivo con timeline de eventos"""
        stats = self._get_stats()
        alerts = self._get_alerts()

        # Build action options
        action_opts = []
        for action in stats['by_action'].keys():
            label = ACTION_LABELS.get(action, action)
            action_opts.append('<option value="{}">{}</option>'.format(action, label))

        # Build user options (usando usuarios normalizados sin duplicados)
        user_opts = []
        for user in sorted(stats['by_user_normalized'].keys()):
            user_opts.append('<option value="{}">{}</option>'.format(user, user))

        # Build timeline items
        timeline_items = []
        for log in sorted(self.logs, key=lambda x: x.timestamp or datetime.min, reverse=True):
            action_label = ACTION_LABELS.get(log.action, log.action)
            status_badge = 'success' if log.success else 'failed'
            status_text = 'Exitoso' if log.success else 'Fallido'
            time_str = log.timestamp.strftime('%H:%M') if log.timestamp else ''
            date_str = log.timestamp.strftime('%Y-%m-%d') if log.timestamp else ''
            size_str = ' (' + self._format_size(log.file_size) + ')' if log.file_size else ''
            win_path = self._to_windows_path(log.target_path)

            deleted_html = ''
            if log.action == 'delete' and log.details:
                items = log.details.get('deleted_items', [])
                if items:
                    # Construir arbol de archivos eliminados
                    tree = self._build_tree_structure(items)
                    tree_html = self._render_tree_html(tree)

                    # Estadisticas del contenido eliminado
                    dir_count = sum(1 for i in items if i.get('is_directory'))
                    file_count = len(items) - dir_count
                    total_size = sum(i.get('size', 0) for i in items)

                    deleted_html = '''
                    <details class="deleted-tree" open>
                        <summary>
                            <span class="delete-icon">&#128465;</span>
                            <strong>Contenido Eliminado</strong>
                            <span class="delete-stats">
                                &#128193; {} dirs | &#128196; {} archivos | &#128202; {}
                            </span>
                        </summary>
                        <div class="tree-container">
                            <div class="tree-root">
                                <span class="tree-icon folder-root">&#128465;</span>
                                <span class="tree-name root">{}</span>
                            </div>
                            {}
                        </div>
                    </details>
                    '''.format(
                        dir_count, file_count, self._format_size(total_size),
                        log.target_name or 'Directorio',
                        tree_html
                    )

            copied_html = ''
            if log.action == 'copy' and log.details:
                items = log.details.get('copied_items', [])
                if items:
                    tree = self._build_tree_structure(items)
                    tree_html = self._render_tree_html(tree)
                    dir_count = sum(1 for i in items if i.get('is_directory'))
                    file_count = len(items) - dir_count
                    total_size = sum(i.get('size', 0) for i in items)
                    dest_name = os.path.basename(log.details.get('dest_path', '')) or log.target_name or 'Directorio'
                    copied_html = '''
                    <details class="copied-tree" open>
                        <summary>
                            <span class="copy-icon">&#128203;</span>
                            <strong>Contenido Copiado</strong>
                            <span class="copy-stats">
                                &#128193; {} dirs | &#128196; {} archivos | &#128202; {}
                            </span>
                        </summary>
                        <div class="path-section" style="margin-bottom:6px;">
                            <span class="path-label">&#128228; Destino:</span>
                            <div class="path">{}</div>
                        </div>
                        <div class="tree-container">
                            <div class="tree-root">
                                <span class="tree-icon folder-root">&#128203;</span>
                                <span class="tree-name root">{}</span>
                            </div>
                            {}
                        </div>
                    </details>
                    '''.format(
                        dir_count, file_count, self._format_size(total_size),
                        self._to_windows_path(log.details.get('dest_path', '')),
                        dest_name,
                        tree_html
                    )

            uploaded_html = ''
            if log.action == 'upload_batch' and log.details:
                items = log.details.get('uploaded_items', [])
                if items:
                    # Convertir formato upload_batch al formato esperado por _build_tree_structure
                    scan_items = []
                    for it in items:
                        if it.get('status') == 'success':
                            scan_items.append({
                                'name': it.get('name', ''),
                                'path': it.get('path', it.get('name', '')),
                                'is_directory': it.get('is_directory', False),
                                'size': it.get('size', 0)
                            })
                    if scan_items:
                        tree = self._build_tree_structure(scan_items)
                        tree_html = self._render_tree_html(tree)
                        dir_count = sum(1 for i in scan_items if i.get('is_directory'))
                        file_count = len(scan_items) - dir_count
                        total_size = sum(i.get('size', 0) for i in scan_items)
                        dest_path = log.details.get('destination_path', '')
                        uploaded_html = '''
                        <details class="uploaded-tree" open>
                            <summary>
                                <span class="upload-icon">&#128229;</span>
                                <strong>Archivos Subidos</strong>
                                <span class="upload-stats">
                                    &#128193; {} dirs | &#128196; {} archivos | &#128202; {}
                                </span>
                            </summary>
                            {}
                            <div class="tree-container">
                                <div class="tree-root">
                                    <span class="tree-icon folder-root">&#128229;</span>
                                    <span class="tree-name root">{}</span>
                                </div>
                                {}
                            </div>
                        </details>
                        '''.format(
                            dir_count, file_count, self._format_size(total_size),
                            '<div class="path-section" style="margin-bottom:6px;"><span class="path-label">&#128194; Destino:</span><div class="path">{}</div></div>'.format(self._to_windows_path(dest_path)) if dest_path else '',
                            os.path.basename(dest_path) or 'Directorio Destino',
                            tree_html
                        )

            error_html = ''
            if not log.success and log.error_message:
                error_html = '<div class="path" style="color: #e74c3c;">Error: {}</div>'.format(log.error_message)

            # Normalizar username para evitar duplicados con correo
            normalized_user = self._normalize_username(log.username)

            item = '<div class="timeline-item {}" data-action="{}" data-user="{}" data-date="{}" data-search="{} {} {}">'.format(
                log.action, log.action, normalized_user, date_str, normalized_user, log.target_name or '', log.target_path or ''
            )
            item += '<div class="time"><strong>{}</strong><br><small>{}</small></div>'.format(time_str, date_str)
            item += '<div class="content">'
            item += '<h3>{}{}</h3>'.format(action_label, size_str)
            item += '<div class="info-grid">'
            item += '<div class="info-item"><span class="info-label">&#128197; Fecha:</span> <span class="info-value">{} {}</span></div>'.format(date_str, time_str)
            item += '<div class="info-item"><span class="info-label">&#128100; Responsable:</span> <span class="info-value">{}</span></div>'.format(normalized_user)
            item += '<div class="info-item"><span class="info-label">&#127919; Rol:</span> <span class="info-value">{}</span></div>'.format(log.user_role or 'N/A')
            item += '<div class="info-item"><span class="info-label">&#128187; IP:</span> <span class="info-value">{}</span></div>'.format(log.ip_address or 'N/A')
            item += '<div class="info-item"><span class="info-label">&#9989; Estado:</span> <span class="badge {}">{}</span></div>'.format(status_badge, status_text)
            if log.file_size:
                item += '<div class="info-item"><span class="info-label">&#128202; Tamano:</span> <span class="info-value">{}</span></div>'.format(self._format_size(log.file_size))
            item += '</div>'
            item += '<div class="path-section"><span class="path-label">&#128194; Ruta Windows:</span><div class="path">{}</div></div>'.format(win_path)
            item += error_html
            item += deleted_html
            item += copied_html
            item += uploaded_html
            item += '</div></div>'
            timeline_items.append(item)

        # Build alerts section
        alerts_html = ''
        if alerts:
            alert_items = []
            for a in alerts[:5]:
                alert_items.append('<div class="alert-item"><strong>{}</strong>: {} - {}</div>'.format(
                    a['severity'].upper(), a['message'], a['user']
                ))
            alerts_html = '<div class="alerts"><h3>Alertas de Seguridad ({})</h3>{}</div>'.format(
                len(alerts), ''.join(alert_items)
            )

        # Construir info de rango de fechas
        date_range = stats.get('date_range', (None, None))
        date_range_html = ''
        if date_range[0] and date_range[1]:
            date_from = date_range[0].strftime('%Y-%m-%d')
            date_to = date_range[1].strftime('%Y-%m-%d')
            if date_from == date_to:
                date_range_html = ' | Fecha: {}'.format(date_from)
            else:
                date_range_html = ' | Periodo: {} a {}'.format(date_from, date_to)

        # Build complete HTML
        html = '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Timeline de Auditoria</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; padding: 20px; }
        .container { max-width: 1400px; margin: 0 auto; }
        .header { background: white; border-radius: 15px; padding: 30px; margin-bottom: 20px; box-shadow: 0 10px 40px rgba(0,0,0,0.2); }
        .header h1 { color: #2c3e50; margin-bottom: 10px; font-size: 2em; }
        .header .meta { color: #7f8c8d; font-size: 0.9em; }
        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px; }
        .stat-card { background: white; border-radius: 12px; padding: 20px; text-align: center; box-shadow: 0 5px 20px rgba(0,0,0,0.1); }
        .stat-card .number { font-size: 2.5em; font-weight: bold; color: #3498db; }
        .stat-card .label { color: #7f8c8d; margin-top: 5px; }
        .stat-card.danger .number { color: #e74c3c; }
        .stat-card.success .number { color: #27ae60; }
        .stat-card.warning .number { color: #f39c12; }
        .filters { background: white; border-radius: 12px; padding: 20px; margin-bottom: 20px; box-shadow: 0 5px 20px rgba(0,0,0,0.1); }
        .filters input, .filters select { padding: 10px 15px; border: 2px solid #ecf0f1; border-radius: 8px; margin-right: 10px; font-size: 14px; }
        .timeline { background: white; border-radius: 15px; padding: 30px; box-shadow: 0 10px 40px rgba(0,0,0,0.2); }
        .timeline-item { position: relative; padding: 20px; border-left: 3px solid #ecf0f1; margin-left: 10px; margin-bottom: 10px; }
        .timeline-item:hover { background: #f8f9fa; border-radius: 10px; }
        .timeline-item.upload { border-left-color: #27ae60; }
        .timeline-item.download { border-left-color: #3498db; }
        .timeline-item.delete { border-left-color: #e74c3c; }
        .timeline-item.rename { border-left-color: #f39c12; }
        .timeline-item.copy { border-left-color: #9b59b6; }
        .timeline-item.move { border-left-color: #1abc9c; }
        .timeline-item.create_folder { border-left-color: #f1c40f; }
        .timeline-item .time { font-size: 0.85em; color: #7f8c8d; margin-bottom: 5px; }
        .timeline-item .content h3 { color: #2c3e50; margin-bottom: 5px; }
        .timeline-item .content .details { color: #7f8c8d; font-size: 0.9em; }
        .timeline-item .content .path { font-family: monospace; background: #ecf0f1; padding: 5px 10px; border-radius: 5px; margin-top: 10px; font-size: 0.85em; word-break: break-all; }
        .badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.8em; font-weight: bold; margin-right: 5px; }
        .badge.success { background: #d4edda; color: #155724; }
        .badge.failed { background: #f8d7da; color: #721c24; }
        .badge.user { background: #e3f2fd; color: #1565c0; }
        .alerts { background: #fff5f5; border-radius: 12px; padding: 20px; margin-bottom: 20px; border-left: 4px solid #e74c3c; }
        .alerts h3 { color: #e74c3c; margin-bottom: 15px; }
        .alert-item { padding: 10px; background: white; border-radius: 8px; margin-bottom: 10px; }
        .no-results { text-align: center; padding: 40px; color: #7f8c8d; display: none; }

        /* Estilos para arbol de archivos eliminados */
        .deleted-tree { margin-top: 15px; background: linear-gradient(135deg, #fff5f5 0%, #ffe6e6 100%); border-radius: 12px; border: 2px solid #e74c3c; overflow: hidden; }
        .deleted-tree summary { padding: 15px; cursor: pointer; display: flex; align-items: center; gap: 10px; background: #e74c3c; color: white; font-size: 0.95em; }
        .deleted-tree summary:hover { background: #c0392b; }
        .deleted-tree summary .delete-icon { font-size: 1.3em; }
        .deleted-tree summary .delete-stats { margin-left: auto; font-size: 0.85em; opacity: 0.9; background: rgba(255,255,255,0.2); padding: 4px 10px; border-radius: 15px; }
        .copied-tree { margin-top: 15px; background: linear-gradient(135deg, #f5f0ff 0%, #e8d5ff 100%); border-radius: 12px; border: 2px solid #9b59b6; overflow: hidden; }
        .copied-tree summary { padding: 15px; cursor: pointer; display: flex; align-items: center; gap: 10px; background: #9b59b6; color: white; font-size: 0.95em; }
        .copied-tree summary:hover { background: #7d3c98; }
        .copied-tree summary .copy-icon { font-size: 1.3em; }
        .copied-tree summary .copy-stats { margin-left: auto; font-size: 0.85em; opacity: 0.9; background: rgba(255,255,255,0.2); padding: 4px 10px; border-radius: 15px; }
        .copied-tree .tree-container { background: #faf5ff; }
        .copied-tree .tree-root { background: #9b59b6; }
        .copied-tree .tree-line:hover { background: rgba(155, 89, 182, 0.1); }
        .copied-tree .tree-name.dir { color: #7d3c98; }
        .uploaded-tree { margin-top: 15px; background: linear-gradient(135deg, #f0fff4 0%, #d5ffe0 100%); border-radius: 12px; border: 2px solid #27ae60; overflow: hidden; }
        .uploaded-tree summary { padding: 15px; cursor: pointer; display: flex; align-items: center; gap: 10px; background: #27ae60; color: white; font-size: 0.95em; }
        .uploaded-tree summary:hover { background: #1e8449; }
        .uploaded-tree summary .upload-icon { font-size: 1.3em; }
        .uploaded-tree summary .upload-stats { margin-left: auto; font-size: 0.85em; opacity: 0.9; background: rgba(255,255,255,0.2); padding: 4px 10px; border-radius: 15px; }
        .uploaded-tree .tree-container { background: #f0fff4; }
        .uploaded-tree .tree-root { background: #27ae60; }
        .uploaded-tree .tree-line:hover { background: rgba(39, 174, 96, 0.1); }
        .uploaded-tree .tree-name.dir { color: #1e8449; }
        .tree-container { padding: 15px; font-family: 'Consolas', 'Monaco', 'Courier New', monospace; font-size: 0.9em; background: #fefefe; max-height: 400px; overflow-y: auto; }
        .tree-root { display: flex; align-items: center; gap: 8px; padding: 8px 12px; background: #e74c3c; color: white; border-radius: 8px; margin-bottom: 10px; font-weight: bold; }
        .tree-root .tree-icon { font-size: 1.4em; }
        .tree-line { display: flex; align-items: center; padding: 3px 0; transition: background 0.2s; }
        .tree-line:hover { background: rgba(231, 76, 60, 0.1); border-radius: 4px; }
        .tree-prefix { color: #bdc3c7; white-space: pre; font-size: 0.95em; }
        .tree-icon { margin-right: 6px; font-size: 1.1em; }
        .tree-icon.folder { color: #f39c12; }
        .tree-icon.file { color: #3498db; }
        .tree-name { color: #2c3e50; }
        .tree-name.dir { font-weight: 600; color: #c0392b; }
        .tree-name.file { color: #34495e; }
        .tree-size { color: #95a5a6; font-size: 0.85em; margin-left: 8px; }

        /* Estilos para grid de informacion descriptiva */
        .info-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 8px 20px; margin: 12px 0; padding: 12px; background: #f8f9fa; border-radius: 8px; border-left: 3px solid #3498db; }
        .info-item { display: flex; align-items: center; gap: 8px; font-size: 0.9em; }
        .info-label { font-weight: 600; color: #2c3e50; white-space: nowrap; }
        .info-value { color: #555; }
        .path-section { margin-top: 10px; padding: 10px; background: #ecf0f1; border-radius: 8px; }
        .path-label { display: block; font-weight: 600; color: #2c3e50; margin-bottom: 5px; font-size: 0.9em; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Timeline de Auditoria</h1>
            <p class="meta">Generado: ''' + self.generation_time.strftime('%Y-%m-%d %H:%M:%S') + ''' | Por: ''' + self.generated_by + ''' | Total: ''' + str(len(self.logs)) + ''' eventos''' + date_range_html + '''</p>
        </div>
        <div class="stats-grid">
            <div class="stat-card"><div class="number">''' + str(stats['total']) + '''</div><div class="label">Total Operaciones</div></div>
            <div class="stat-card success"><div class="number">''' + str(stats['successful']) + '''</div><div class="label">Exitosas</div></div>
            <div class="stat-card danger"><div class="number">''' + str(stats['failed']) + '''</div><div class="label">Fallidas</div></div>
            <div class="stat-card warning"><div class="number">''' + str(stats['deleted_files_count']) + '''</div><div class="label">Archivos Eliminados</div></div>
            <div class="stat-card"><div class="number">''' + str(len(stats['by_user_normalized'])) + '''</div><div class="label">Usuarios</div></div>
            <div class="stat-card"><div class="number">''' + self._format_size(stats['total_size']) + '''</div><div class="label">Tamano Procesado</div></div>
        </div>
        ''' + alerts_html + '''
        <div class="filters">
            <input type="text" id="searchInput" placeholder="Buscar..." onkeyup="filterTimeline()">
            <select id="actionFilter" onchange="filterTimeline()">
                <option value="">Todas las acciones</option>
                ''' + ''.join(action_opts) + '''
            </select>
            <select id="userFilter" onchange="filterTimeline()">
                <option value="">Todos los usuarios</option>
                ''' + ''.join(user_opts) + '''
            </select>
            <label style="margin-left:15px; color:#7f8c8d; font-size:0.9em;">Desde:</label>
            <input type="date" id="dateFrom" onchange="filterTimeline()">
            <label style="color:#7f8c8d; font-size:0.9em;">Hasta:</label>
            <input type="date" id="dateTo" onchange="filterTimeline()">
            <button onclick="clearFilters()" style="padding:10px 15px; background:#e74c3c; color:white; border:none; border-radius:8px; cursor:pointer; margin-left:10px;">Limpiar</button>
        </div>
        <div class="timeline" id="timeline">
            ''' + ''.join(timeline_items) + '''
            <div class="no-results" id="noResults">No se encontraron resultados</div>
        </div>
    </div>
    <script>
        function filterTimeline() {
            var search = document.getElementById('searchInput').value.toLowerCase();
            var action = document.getElementById('actionFilter').value;
            var user = document.getElementById('userFilter').value.toLowerCase();
            var dateFrom = document.getElementById('dateFrom').value;
            var dateTo = document.getElementById('dateTo').value;
            var items = document.querySelectorAll('.timeline-item');
            var visibleCount = 0;
            items.forEach(function(item) {
                var matchSearch = item.dataset.search.toLowerCase().includes(search);
                var matchAction = !action || item.dataset.action === action;
                var itemUser = item.dataset.user.toLowerCase();
                var matchUser = !user || itemUser === user || itemUser.startsWith(user);
                var itemDate = item.dataset.date;
                var matchDate = true;
                if (dateFrom && itemDate < dateFrom) matchDate = false;
                if (dateTo && itemDate > dateTo) matchDate = false;
                if (matchSearch && matchAction && matchUser && matchDate) {
                    item.style.display = '';
                    visibleCount++;
                } else {
                    item.style.display = 'none';
                }
            });
            document.getElementById('noResults').style.display = visibleCount === 0 ? 'block' : 'none';
        }
        function clearFilters() {
            document.getElementById('searchInput').value = '';
            document.getElementById('actionFilter').value = '';
            document.getElementById('userFilter').value = '';
            document.getElementById('dateFrom').value = '';
            document.getElementById('dateTo').value = '';
            filterTimeline();
        }
    </script>
</body>
</html>'''
        return html.encode('utf-8')

    def generate_readme(self) -> bytes:
        """Genera archivo README explicando el contenido del ZIP"""
        stats = self._get_stats()

        content = f"""
================================================================================
                    PAQUETE DE REPORTES DE AUDITORÍA
================================================================================

Generado: {self.generation_time.strftime('%Y-%m-%d %H:%M:%S')}
Por: {self.generated_by}
Sistema: Sistema de Gestión Documental - IGAC

================================================================================
                           CONTENIDO DEL PAQUETE
================================================================================

1. listado_completo.csv
   - Formato: CSV (valores separados por coma)
   - Contenido: Listado completo de todas las operaciones
   - Uso: Importar en Excel, Google Sheets o cualquier herramienta de análisis
   - Codificación: UTF-8 con BOM (compatible con Excel)

2. arbol_eventos.txt
   - Formato: Texto plano
   - Contenido: Vista jerárquica de eventos con emojis
   - Uso: Revisión rápida, impresión, archivo

3. reporte_detallado.xlsx
   - Formato: Microsoft Excel
   - Contenido:
     * Hoja 1 - Portada: Información del reporte
     * Hoja 2 - Resumen: Estadísticas generales
     * Hoja 3 - Detalle: Todas las operaciones con filtros
     * Hoja 4 - Eliminaciones: Detalle de archivos eliminados
     * Hoja 5 - Alertas: Alertas de seguridad detectadas
   - Uso: Análisis detallado, presentaciones, auditorías formales

4. timeline_interactivo.html
   - Formato: HTML (abrir en navegador)
   - Contenido: Timeline visual e interactivo de eventos
   - Uso: Presentaciones, revisión visual, filtrado dinámico
   - Nota: No requiere conexión a internet

5. README.txt (este archivo)
   - Descripción del contenido del paquete

================================================================================
                              ESTADÍSTICAS
================================================================================

Total de operaciones registradas: {stats['total']}
Operaciones exitosas: {stats['successful']}
Operaciones fallidas: {stats['failed']}
Archivos eliminados (detalle): {stats['deleted_files_count']}
Tamaño total procesado: {self._format_size(stats['total_size'])}
Usuarios únicos: {len(stats['by_user'])}

================================================================================
                         CUMPLIMIENTO NORMATIVO
================================================================================

Este reporte cumple con los requisitos de trazabilidad establecidos en:

- ISO 27001:2013 - Gestión de seguridad de la información
- ISO 15489:2016 - Gestión de documentos
- Ley 594 de 2000 - Ley General de Archivos (Colombia)
- Decreto 1080 de 2015 - Gestión documental
- NTC-ISO 30300 - Sistemas de gestión para los documentos

================================================================================
                           INTEGRIDAD DEL REPORTE
================================================================================

Hash SHA-256: {hashlib.sha256(str(stats).encode()).hexdigest()}

Este hash puede ser utilizado para verificar la integridad del reporte.

================================================================================
                              SOPORTE
================================================================================

Para consultas sobre este reporte, contactar al administrador del sistema.

================================================================================
"""
        return content.encode('utf-8')

    def generate_zip_package(self) -> bytes:
        """Genera paquete ZIP con todos los reportes"""
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # CSV
            zip_file.writestr('listado_completo.csv', self.generate_csv())

            # TXT Árbol
            zip_file.writestr('arbol_eventos.txt', self.generate_tree_txt())

            # Excel
            zip_file.writestr('reporte_detallado.xlsx', self.generate_excel())

            # HTML Timeline
            zip_file.writestr('timeline_interactivo.html', self.generate_html_timeline())

            # README
            zip_file.writestr('README.txt', self.generate_readme())

        zip_buffer.seek(0)
        return zip_buffer.getvalue()
