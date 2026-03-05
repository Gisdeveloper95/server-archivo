"""
Vistas de administración de usuarios (solo para superadmins)
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.utils import timezone
from .models import User, UserPermission
from .serializers import UserSerializer, UserPermissionSerializer, BulkPermissionAssignmentSerializer
import secrets
import string
from django.db import transaction
from datetime import datetime


class AdminUserViewSet(viewsets.ViewSet):
    """
    ViewSet para gestión de usuarios por parte de superadmins
    """
    permission_classes = [IsAuthenticated]

    def _check_superadmin(self, user):
        """Verificar que el usuario es superadmin"""
        if user.role != 'superadmin':
            return Response(
                {'error': 'Solo superadmins pueden acceder a esta función'},
                status=status.HTTP_403_FORBIDDEN
            )
        return None

    def list(self, request):
        """
        GET /api/admin/users
        Lista todos los usuarios del sistema

        Query params:
            - search: Buscar por username, email, first_name, last_name
            - role: Filtrar por rol
            - per_page: Límite de resultados (default: todos)
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        users = User.objects.all().order_by('-created_at')

        # Filtro de búsqueda
        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            users = users.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )

        # Filtro por rol
        role = request.query_params.get('role')
        if role:
            users = users.filter(role=role)

        # Límite de resultados
        per_page = request.query_params.get('per_page')
        if per_page:
            try:
                users = users[:int(per_page)]
            except ValueError:
                pass

        serializer = UserSerializer(users, many=True)

        return Response({
            'users': serializer.data,
            'results': serializer.data,  # Compatibilidad con frontend
            'total': users.count() if not per_page else User.objects.count()
        })

    def create(self, request):
        """
        POST /api/admin/users
        Crea un nuevo usuario

        Body:
        {
            "username": "string",
            "email": "string@igac.gov.co",
            "first_name": "string",
            "last_name": "string",
            "role": "consultation|consultation_edit|admin|superadmin",
            "password": "string (12 chars min)",
            "send_email": boolean
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        data = request.data

        # Validaciones
        if not data.get('email', '').endswith('@igac.gov.co'):
            return Response(
                {'error': 'El email debe ser corporativo del IGAC (@igac.gov.co)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if len(data.get('password', '')) < 12:
            return Response(
                {'error': 'La contraseña debe tener al menos 12 caracteres'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si el username ya existe
        if User.objects.filter(username=data.get('username')).exists():
            return Response(
                {'error': 'El username ya está en uso'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si el email ya existe
        if User.objects.filter(email=data.get('email')).exists():
            return Response(
                {'error': 'El email ya está registrado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear usuario
        try:
            user = User.objects.create(
                username=data['username'],
                email=data['email'],
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                role=data.get('role', 'consultation'),
                password=make_password(data['password']),
                is_active=True
            )

            # Enviar email si se solicitó
            if data.get('send_email', False):
                try:
                    print(f"[DEBUG] Intentando enviar email a {user.email}...")
                    self._send_welcome_email(user, data['password'])
                    print(f"[DEBUG] Email enviado exitosamente a {user.email}")
                except Exception as e:
                    print(f"[ERROR] Error enviando email: {type(e).__name__}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    # No fallar la creación si el email falla

            serializer = UserSerializer(user)
            return Response({
                'message': 'Usuario creado exitosamente',
                'user': serializer.data
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Error al crear usuario: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def partial_update(self, request, pk=None):
        """
        PATCH /api/admin/users/{id}
        Actualiza un usuario existente

        Body:
        {
            "first_name": "string",
            "last_name": "string",
            "role": "consultation|consultation_edit|admin|superadmin",
            "is_active": boolean
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # No permitir que un superadmin se desactive a sí mismo
        if user.id == request.user.id and request.data.get('is_active') == False:
            return Response(
                {'error': 'No puedes desactivarte a ti mismo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Actualizar campos permitidos
        updated_fields = []

        if 'first_name' in request.data:
            user.first_name = request.data['first_name']
            updated_fields.append('first_name')
        if 'last_name' in request.data:
            user.last_name = request.data['last_name']
            updated_fields.append('last_name')
        if 'role' in request.data:
            user.role = request.data['role']
            updated_fields.append('role')
        if 'is_active' in request.data:
            user.is_active = request.data['is_active']
            updated_fields.append('is_active')

        # Solo guardar los campos que se actualizaron
        if updated_fields:
            user.save(update_fields=updated_fields)

        serializer = UserSerializer(user)
        return Response({
            'message': 'Usuario actualizado exitosamente',
            'user': serializer.data
        })

    def destroy(self, request, pk=None):
        """
        DELETE /api/admin/users/{id}
        Elimina un usuario
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # No permitir que un superadmin se elimine a sí mismo
        if user.id == request.user.id:
            return Response(
                {'error': 'No puedes eliminarte a ti mismo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        username = user.username
        user.delete()

        return Response({
            'success': True,
            'message': f'Usuario "{username}" eliminado exitosamente'
        })

    @action(detail=True, methods=['post'], url_path='resend-credentials')
    def resend_credentials(self, request, pk=None):
        """
        POST /api/admin/users/{id}/resend-credentials
        Genera nueva contraseña y la envía por email
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Generar nueva contraseña
        new_password = self._generate_secure_password()
        user.password = make_password(new_password)
        user.save()

        # Enviar email
        try:
            self._send_welcome_email(user, new_password, is_reset=True)
            return Response({
                'message': f'Credenciales enviadas a {user.email}'
            })
        except Exception as e:
            return Response(
                {'error': f'Error al enviar email: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def permissions(self, request, pk=None):
        """
        GET /api/admin/users/{id}/permissions
        Obtiene todos los permisos de un usuario CON TODOS LOS CAMPOS GRANULARES
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        permissions = UserPermission.objects.filter(
            user=user,
            is_active=True
        ).select_related('granted_by').order_by('base_path')

        # Usar el serializer completo para incluir TODOS los campos
        serializer = UserPermissionSerializer(permissions, many=True)

        return Response({
            'permissions': serializer.data,
            'total': len(serializer.data)
        })

    @action(detail=True, methods=['post'], url_path='assign-permission')
    def assign_permission(self, request, pk=None):
        """
        POST /api/admin/users/{id}/assign-permission
        Asigna un nuevo permiso a un usuario

        Body:
        {
            "base_path": "string",
            "can_read": boolean,
            "can_write": boolean,
            "can_delete": boolean,
            "exempt_from_dictionary": boolean
        }
        """
        import traceback
        print(f"[DEBUG] assign_permission llamado para user_id={pk}")
        print(f"[DEBUG] Request data: {request.data}")

        try:
            error = self._check_superadmin(request.user)
            if error:
                print(f"[DEBUG] Error de superadmin check")
                return error

            try:
                user = User.objects.get(pk=pk)
                print(f"[DEBUG] Usuario encontrado: {user.username}")
            except User.DoesNotExist:
                return Response(
                    {'error': 'Usuario no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = request.data
            base_path = data.get('base_path', '').strip()
        except Exception as e:
            print(f"[ERROR] Exception en assign_permission inicio: {str(e)}")
            traceback.print_exc()
            return Response(
                {'error': f'Error interno: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Envolver todo el resto en try-except para debugging
        try:
            from services.path_converter import PathConverter

            # Si es ruta Windows UNC, convertir a ruta relativa
            if PathConverter.is_windows_path(base_path):
                linux_full_path = PathConverter.windows_to_linux(base_path)
                if linux_full_path.startswith(PathConverter.LINUX_MOUNT_POINT):
                    base_path = linux_full_path[len(PathConverter.LINUX_MOUNT_POINT):].lstrip('/')
                else:
                    base_path = linux_full_path.lstrip('/')
                print(f"[DEBUG] Ruta Windows UNC convertida a relativa: '{base_path}'")
            else:
                # Normalizar la ruta: convertir backslashes a forward slashes y limpiar
                base_path = base_path.replace('\\', '/')
                # Remover prefijo Windows si fue pegado parcialmente (sin \\)
                prefix_to_remove = 'repositorio/DirGesCat/2510SP/H_Informacion_Consulta/Sub_Proy/'
                if prefix_to_remove.lower() in base_path.lower():
                    idx = base_path.lower().find(prefix_to_remove.lower())
                    base_path = base_path[idx + len(prefix_to_remove):]

            # Remover slashes al inicio y final
            base_path = base_path.strip('/')

            # Permitir base_path vacío (significa acceso a toda la carpeta Sub_Proy)
            # Si está vacío, es válido y se almacenará como cadena vacía

            # Validación adicional: verificar que no contenga caracteres peligrosos (solo si no está vacío)
            if base_path and ('..' in base_path or '//' in base_path):
                return Response(
                    {'error': 'La ruta contiene secuencias no permitidas (.., //)'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verificar si ya existe un permiso para esta ruta (activo o inactivo)
            existing_active = UserPermission.objects.filter(
                user=user,
                base_path=base_path,
                is_active=True
            ).first()

            if existing_active:
                return Response(
                    {'error': 'El usuario ya tiene permisos activos para esta ruta'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Si existe un permiso inactivo para esta ruta, eliminarlo para poder crear uno nuevo
            # (evita violación de constraint unique en user_id + base_path)
            existing_inactive = UserPermission.objects.filter(
                user=user,
                base_path=base_path,
                is_active=False
            ).first()

            if existing_inactive:
                print(f"[DEBUG] Eliminando permiso inactivo existente ID={existing_inactive.id} para reasignar")
                existing_inactive.delete()

            # Procesar fecha de vencimiento
            from django.utils.dateparse import parse_datetime
            from datetime import datetime

            expires_at = data.get('expires_at')
            if expires_at:
                # Si viene como string ISO, parsear
                if isinstance(expires_at, str):
                    expires_at = parse_datetime(expires_at)
                    # Asegurar que sea timezone-aware
                    if expires_at and timezone.is_naive(expires_at):
                        expires_at = timezone.make_aware(expires_at)

                # Si no se especificó, usar 31 de diciembre del año actual
                if not expires_at:
                    current_year = timezone.now().year
                    naive_dt = datetime(current_year, 12, 31, 23, 59, 59)
                    expires_at = timezone.make_aware(naive_dt)
            else:
                # Valor por defecto: 31 de diciembre del año actual
                current_year = timezone.now().year
                naive_dt = datetime(current_year, 12, 31, 23, 59, 59)
                expires_at = timezone.make_aware(naive_dt)

            print(f"[DEBUG] Creando permiso con base_path='{base_path}', expires_at={expires_at}")

            # Crear permiso
            permission = UserPermission.objects.create(
                user=user,
                base_path=base_path,
                can_read=data.get('can_read', True),
                can_write=data.get('can_write', False),
                can_delete=data.get('can_delete', False),
                can_create_directories=data.get('can_create_directories', True),
                exempt_from_dictionary=data.get('exempt_from_dictionary', False),
                edit_permission_level=data.get('edit_permission_level'),
                inheritance_mode=data.get('inheritance_mode', 'total'),
                blocked_paths=data.get('blocked_paths', []),
                read_only_paths=data.get('read_only_paths', []),
                max_depth=data.get('max_depth'),
                expires_at=expires_at,
                granted_by=request.user,
                notes=data.get('notes', ''),
                authorized_by_email=data.get('authorized_by_email'),
                authorized_by_name=data.get('authorized_by_name')
            )

            print(f"[DEBUG] Permiso creado con ID={permission.id}")

            # Enviar email de notificación al usuario con Excel adjunto
            try:
                from django.core.mail import EmailMessage
                from django.template.loader import render_to_string
                from django.conf import settings

                # Preparar contexto para el template
                base_path_escaped = base_path.replace('/', '\\\\') if base_path else ''
                context = {
                    'user_name': user.get_full_name(),
                    'base_path_display': f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if base_path else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy",
                    'can_read': permission.can_read,
                    'can_write': permission.can_write,
                    'can_delete': permission.can_delete,
                    'can_create_directories': permission.can_create_directories,
                    'exempt_from_dictionary': permission.exempt_from_dictionary,
                    'edit_permission_level': permission.edit_permission_level,
                    'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(permission.edit_permission_level, 'N/A') if permission.edit_permission_level else None,
                    'inheritance_mode': permission.inheritance_mode,
                    'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(permission.inheritance_mode),
                    'granted_at': permission.granted_at.strftime('%d de %B de %Y a las %H:%M'),
                    'expires_at': permission.expires_at.strftime('%d de %B de %Y'),
                    'days_until_expiration': permission.days_until_expiration(),
                    'granted_by': request.user.get_full_name(),
                    'granted_by_role': dict(User.ROLE_CHOICES).get(request.user.role),
                    'frontend_url': settings.FRONTEND_URL,
                    'notes': permission.notes,
                }

                # Renderizar HTML
                html_content = render_to_string('emails/permission_assigned.html', context)

                # Generar Excel
                excel_file = self._generate_excel_report([permission], permission.authorized_by_name or 'Sistema')

                # Preparar nombre de archivo
                timestamp = datetime.now().strftime('%Y%m%d')
                authorized_name = permission.authorized_by_name.replace(' ', '_') if permission.authorized_by_name else 'Sistema'
                filename = f'{timestamp}_Permisos_Repositorio_{authorized_name}.xlsx'

                # Crear email con adjunto
                email = EmailMessage(
                    subject=f'🔐 Permisos de Acceso Asignados - Sistema IGAC',
                    body=f'Se le han asignado permisos de acceso. Vea los detalles en: {settings.FRONTEND_URL}',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email],
                )
                email.content_subtype = 'html'
                email.body = html_content

                # Adjuntar Excel
                email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                # Agregar CC si existe quien autorizó
                if permission.authorized_by_email:
                    email.cc = [permission.authorized_by_email]

                # Enviar
                email.send(fail_silently=False)
                print(f"[OK] Email con Excel enviado exitosamente a {user.email}" +
                      (f" (CC: {permission.authorized_by_email})" if permission.authorized_by_email else ""))
            except Exception as e:
                # Log el error pero no fallar la operación
                print(f"[ERROR] Error enviando email de asignación de permiso a {user.email}: {str(e)}")
                import traceback
                traceback.print_exc()

            return Response({
                'message': 'Permiso asignado exitosamente',
                'permission': {
                    'id': permission.id,
                    'base_path': permission.base_path,
                    'can_read': permission.can_read,
                    'can_write': permission.can_write,
                    'can_delete': permission.can_delete,
                    'exempt_from_dictionary': permission.exempt_from_dictionary,
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            import traceback
            print(f"[ERROR] Exception en assign_permission: {str(e)}")
            traceback.print_exc()
            return Response(
                {'error': f'Error interno: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='assign-multiple-permissions')
    def assign_multiple_permissions(self, request, pk=None):
        """
        POST /api/admin/users/{id}/assign-multiple-permissions
        Asigna múltiples permisos a un usuario y envía un solo correo con todas las rutas

        Body:
        {
            "routes": ["ruta1", "ruta2", "ruta3"],
            "can_read": boolean,
            "can_write": boolean,
            "can_delete": boolean,
            "can_create_directories": boolean,
            "exempt_from_dictionary": boolean,
            "edit_permission_level": string,
            "inheritance_mode": string,
            "blocked_paths": [],
            "read_only_paths": [],
            "max_depth": number | null,
            "expires_at": string,
            "notes": string,
            "authorized_by_email": string,
            "authorized_by_name": string
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data
        routes = data.get('routes', [])

        if not routes or not isinstance(routes, list):
            return Response(
                {'error': 'Se requiere una lista de rutas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Procesar fecha de vencimiento
        from django.utils.dateparse import parse_datetime
        expires_at = data.get('expires_at')
        if expires_at and isinstance(expires_at, str):
            expires_at = parse_datetime(expires_at)
            if expires_at and timezone.is_naive(expires_at):
                expires_at = timezone.make_aware(expires_at)

        created_permissions = []
        errors = []

        # Crear todos los permisos SIN enviar correo todavía
        for base_path in routes:
            try:
                # Normalizar la ruta
                base_path = base_path.replace('\\', '/').strip('/')

                # Validar secuencias peligrosas
                if base_path and ('..' in base_path or '//' in base_path):
                    errors.append(f'{base_path}: Ruta contiene secuencias no permitidas')
                    continue

                # Verificar si ya existe permiso activo
                existing_active = UserPermission.objects.filter(
                    user=user,
                    base_path=base_path,
                    is_active=True
                ).first()

                if existing_active:
                    errors.append(f'{base_path}: Usuario ya tiene permisos activos')
                    continue

                # Eliminar permisos inactivos para esta ruta
                UserPermission.objects.filter(
                    user=user,
                    base_path=base_path,
                    is_active=False
                ).delete()

                # Crear permiso
                permission = UserPermission.objects.create(
                    user=user,
                    base_path=base_path,
                    can_read=data.get('can_read', True),
                    can_write=data.get('can_write', False),
                    can_delete=data.get('can_delete', False),
                    can_create_directories=data.get('can_create_directories', False),
                    exempt_from_dictionary=data.get('exempt_from_dictionary', False),
                    edit_permission_level=data.get('edit_permission_level', 'upload_only'),
                    inheritance_mode=data.get('inheritance_mode', 'total'),
                    blocked_paths=data.get('blocked_paths', []),
                    read_only_paths=data.get('read_only_paths', []),
                    max_depth=data.get('max_depth'),
                    expires_at=expires_at,
                    notes=data.get('notes', ''),
                    granted_by=request.user,
                    authorized_by_email=data.get('authorized_by_email', ''),
                    authorized_by_name=data.get('authorized_by_name', ''),
                    is_active=True
                )
                created_permissions.append(permission)

            except Exception as e:
                errors.append(f'{base_path}: {str(e)}')

        # Si se crearon permisos exitosamente, enviar UN SOLO correo con TODAS las rutas
        if created_permissions:
            try:
                from django.core.mail import EmailMessage
                from django.template.loader import render_to_string
                from datetime import datetime

                # Preparar contexto para el email con TODAS las rutas
                first_perm = created_permissions[0]

                # Construir lista de rutas para mostrar en el email
                routes_list = []
                for perm in created_permissions:
                    base_path_escaped = perm.base_path.replace('/', '\\\\') if perm.base_path else ''
                    full_path = f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if perm.base_path else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy"
                    routes_list.append({
                        'path': full_path,
                        'can_read': perm.can_read,
                        'can_write': perm.can_write,
                        'can_delete': perm.can_delete,
                        'can_create_directories': perm.can_create_directories,
                    })

                context = {
                    'user_name': user.get_full_name(),
                    'routes': routes_list,  # Lista de rutas
                    'route_count': len(created_permissions),
                    'can_read': first_perm.can_read,
                    'can_write': first_perm.can_write,
                    'can_delete': first_perm.can_delete,
                    'can_create_directories': first_perm.can_create_directories,
                    'exempt_from_dictionary': first_perm.exempt_from_dictionary,
                    'edit_permission_level': first_perm.edit_permission_level,
                    'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(first_perm.edit_permission_level, 'N/A') if first_perm.edit_permission_level else None,
                    'inheritance_mode': first_perm.inheritance_mode,
                    'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(first_perm.inheritance_mode),
                    'granted_at': first_perm.granted_at.strftime('%d de %B de %Y a las %H:%M'),
                    'expires_at': first_perm.expires_at.strftime('%d de %B de %Y') if first_perm.expires_at else 'Sin vencimiento',
                    'days_until_expiration': first_perm.days_until_expiration() if first_perm.expires_at else None,
                    'granted_by': request.user.get_full_name(),
                    'granted_by_role': dict(User.ROLE_CHOICES).get(request.user.role),
                    'frontend_url': settings.FRONTEND_URL,
                    'notes': first_perm.notes,
                }

                # Renderizar HTML (usa el mismo template pero con múltiples rutas)
                html_content = render_to_string('emails/permission_assigned.html', context)

                # Generar Excel con TODAS las rutas
                excel_file = self._generate_excel_report(created_permissions, first_perm.authorized_by_name or 'Sistema')

                # Crear nombre de archivo
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                authorized_name = first_perm.authorized_by_name.replace(' ', '_') if first_perm.authorized_by_name else 'Sistema'
                filename = f'{timestamp}_Permisos_Repositorio_{authorized_name}.xlsx'

                # Enviar UN SOLO email con TODAS las rutas
                email = EmailMessage(
                    subject=f'🔐 Permisos de Acceso Asignados ({len(created_permissions)} rutas) - Sistema IGAC',
                    body=html_content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email],
                )
                email.content_subtype = 'html'
                email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                # CC si hay quien autorizó
                if first_perm.authorized_by_email:
                    email.cc = [first_perm.authorized_by_email]

                email.send(fail_silently=False)
                print(f"[OK] Email con {len(created_permissions)} rutas enviado a {user.email}" +
                      (f" (CC: {first_perm.authorized_by_email})" if first_perm.authorized_by_email else ""))

            except Exception as e:
                print(f"[ERROR] Error enviando email consolidado: {str(e)}")
                import traceback
                traceback.print_exc()

        return Response({
            'success': True,
            'created_count': len(created_permissions),
            'error_count': len(errors),
            'errors': errors,
            'permissions': [{
                'id': p.id,
                'base_path': p.base_path,
            } for p in created_permissions]
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['patch'], url_path='permissions/(?P<permission_id>[^/.]+)')
    def update_permission(self, request, pk=None, permission_id=None):
        """
        PATCH /api/admin/permissions/{id}/
        Actualiza un permiso existente

        Body: {
            "can_read": boolean,
            "can_write": boolean,
            "can_delete": boolean,
            "exempt_from_dictionary": boolean,
            "edit_permission_level": "upload_only" | "upload_own" | "upload_all",
            "inheritance_mode": "total" | "blocked" | "limited_depth",
            "blocked_paths": [],
            "read_only_subdirs": [],
            "max_depth": number | null,
            "notes": string,
            "notify_user": boolean,  # Enviar email al usuario
            "notify_leader": boolean  # Enviar email al líder del grupo
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            permission = UserPermission.objects.get(pk=permission_id, is_active=True)
        except UserPermission.DoesNotExist:
            return Response(
                {'error': 'Permiso no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data

        # Extraer flags de notificación
        notify_user = data.get('notify_user', False)
        notify_leader = data.get('notify_leader', False)

        # DEBUG: Log notification flags
        print(f"[DEBUG UPDATE_PERMISSION] Permission ID: {permission_id}")
        print(f"[DEBUG UPDATE_PERMISSION] notify_user = {notify_user}")
        print(f"[DEBUG UPDATE_PERMISSION] notify_leader = {notify_leader}")
        print(f"[DEBUG UPDATE_PERMISSION] User email: {permission.user.email}")
        print(f"[DEBUG UPDATE_PERMISSION] authorized_by_email: {permission.authorized_by_email}")

        # Validar que al menos un permiso básico esté activado
        can_read = data.get('can_read', permission.can_read)
        can_write = data.get('can_write', permission.can_write)
        can_delete = data.get('can_delete', permission.can_delete)

        if not (can_read or can_write or can_delete):
            return Response(
                {'error': 'Debe haber al menos un permiso activado (lectura, escritura o eliminación)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Actualizar campos básicos
        permission.can_read = can_read
        permission.can_write = can_write
        permission.can_delete = can_delete
        permission.exempt_from_dictionary = data.get('exempt_from_dictionary', permission.exempt_from_dictionary)

        # Campos granulares
        if 'edit_permission_level' in data:
            permission.edit_permission_level = data['edit_permission_level']
        if 'inheritance_mode' in data:
            permission.inheritance_mode = data['inheritance_mode']
        if 'blocked_paths' in data:
            # Normalizar rutas bloqueadas (quitar prefijo duplicado)
            blocked_paths_raw = data['blocked_paths']
            blocked_paths_normalized = []
            for path in blocked_paths_raw:
                # Normalizar todas las barras a forward slash
                path_normalized = path.replace('\\', '/')
                # Definir el prefijo base que debe removerse
                prefix_to_remove = 'repositorio/DirGesCat/2510SP/H_Informacion_Consulta/Sub_Proy/'
                # Remover el prefijo si existe
                if prefix_to_remove in path_normalized:
                    idx = path_normalized.find(prefix_to_remove)
                    path_normalized = path_normalized[idx + len(prefix_to_remove):]
                # Limpiar barras iniciales y finales
                path_normalized = path_normalized.strip('/').strip('\\')
                blocked_paths_normalized.append(path_normalized)
            permission.blocked_paths = blocked_paths_normalized
        if 'read_only_paths' in data:
            # Normalizar rutas de solo lectura (quitar prefijo duplicado)
            read_only_paths_raw = data['read_only_paths']
            read_only_paths_normalized = []
            for path in read_only_paths_raw:
                # Normalizar todas las barras a forward slash
                path_normalized = path.replace('\\', '/')
                # Definir el prefijo base que debe removerse
                prefix_to_remove = 'repositorio/DirGesCat/2510SP/H_Informacion_Consulta/Sub_Proy/'
                # Remover el prefijo si existe
                if prefix_to_remove in path_normalized:
                    idx = path_normalized.find(prefix_to_remove)
                    path_normalized = path_normalized[idx + len(prefix_to_remove):]
                # Limpiar barras iniciales y finales
                path_normalized = path_normalized.strip('/').strip('\\')
                read_only_paths_normalized.append(path_normalized)
            permission.read_only_paths = read_only_paths_normalized
        if 'max_depth' in data:
            permission.max_depth = data['max_depth']
        if 'notes' in data:
            permission.notes = data['notes']
        if 'can_create_directories' in data:
            permission.can_create_directories = data['can_create_directories']

        # Procesar fecha de vencimiento
        if 'expires_at' in data:
            expires_at = data.get('expires_at')
            if expires_at:
                from django.utils.dateparse import parse_datetime
                from datetime import datetime

                if isinstance(expires_at, str):
                    # Intentar parsear como fecha ISO
                    expires_at = parse_datetime(expires_at)
                    if not expires_at:
                        # Si falla, intentar parsear solo la fecha (YYYY-MM-DD)
                        try:
                            naive_dt = datetime.strptime(data.get('expires_at'), '%Y-%m-%d')
                            expires_at = timezone.make_aware(naive_dt)
                        except:
                            pass

                if expires_at:
                    permission.expires_at = expires_at
                    # Si estaba vencido y se actualiza la fecha, reactivar el permiso
                    if not permission.is_active:
                        permission.is_active = True
                        permission.revoked_at = None
                    # Resetear flags de notificación para que se envíen nuevamente
                    permission.expiration_notified_7days = False
                    permission.expiration_notified_3days = False

        permission.save()

        # ==================== NOTIFICACIONES POR EMAIL ====================
        user = permission.user
        emails_sent = []

        # 1. ENVIAR EMAIL AL USUARIO (si notify_user=True)
        if notify_user:
            try:
                # Generar Excel del permiso actualizado
                permissions_queryset = UserPermission.objects.filter(pk=permission.id, is_active=True).select_related('user')
                excel_file = self._generate_excel_report(permissions_queryset, permission.authorized_by_name or 'Sistema')

                # Preparar contexto para el template de MODIFICACIÓN
                base_path_escaped = permission.base_path.replace('/', '\\\\') if permission.base_path else ''
                context = {
                    'user_name': user.get_full_name(),
                    'base_path_display': f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if permission.base_path else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy",
                    'blocked_paths': permission.blocked_paths or [],
                    'read_only_paths': permission.read_only_paths or [],
                    'can_read': permission.can_read,
                    'can_write': permission.can_write,
                    'can_delete': permission.can_delete,
                    'can_create_directories': permission.can_create_directories,
                    'exempt_from_dictionary': permission.exempt_from_dictionary,
                    'edit_permission_level': permission.edit_permission_level,
                    'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(permission.edit_permission_level, 'N/A') if permission.edit_permission_level else None,
                    'inheritance_mode': permission.inheritance_mode,
                    'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(permission.inheritance_mode),
                    'modified_at': datetime.now().strftime('%d de %B de %Y a las %H:%M'),
                    'expires_at': permission.expires_at.strftime('%d de %B de %Y') if permission.expires_at else 'Sin vencimiento',
                    'modified_by': f"{request.user.get_full_name()} ({request.user.email})",
                    'authorized_by_name': permission.authorized_by_name or 'Sistema',
                    'frontend_url': settings.FRONTEND_URL,
                    'notes': permission.notes or '',
                }

                # Renderizar HTML con template de MODIFICACIÓN
                html_content = render_to_string('emails/permission_modified.html', context)

                # Preparar archivo adjunto
                timestamp = datetime.now().strftime('%Y%m%d')
                authorized_name = permission.authorized_by_name.replace(' ', '_') if permission.authorized_by_name else 'Sistema'
                filename = f'{timestamp}_Permisos_Modificados_{authorized_name}.xlsx'

                # Enviar email con Excel adjunto
                from django.core.mail import EmailMessage
                email = EmailMessage(
                    subject=f'🔄 Permisos Modificados - Sistema IGAC',
                    body=html_content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email],
                )
                email.content_subtype = 'html'
                email.attach(filename, excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                email.send(fail_silently=False)

                emails_sent.append(user.email)
                print(f"[OK] Email de modificación enviado a {user.email}")

            except Exception as e:
                print(f"[ERROR] Error enviando email al usuario {user.email}: {str(e)}")
                import traceback
                traceback.print_exc()

        # 2. ENVIAR EMAIL AL LÍDER (si notify_leader=True y existe authorized_by_email)
        if notify_leader and permission.authorized_by_email:
            try:
                # Obtener todos los permisos del mismo grupo (si existe)
                if permission.group_name:
                    all_group_permissions = UserPermission.objects.filter(
                        group_name=permission.group_name,
                        is_active=True
                    ).select_related('user').order_by('user__email', 'base_path')

                    # Generar Excel del grupo completo
                    excel_file_leader = self._generate_group_excel_report_by_route(
                        all_group_permissions,
                        permission.authorized_by_name or 'Sistema',
                        permission.group_name
                    )
                else:
                    # Si no hay grupo, enviar solo el permiso individual
                    permissions_queryset = UserPermission.objects.filter(pk=permission.id, is_active=True).select_related('user')
                    excel_file_leader = self._generate_excel_report(permissions_queryset, permission.authorized_by_name or 'Sistema')

                # Preparar contexto para el email del líder
                context_leader = {
                    'leader_name': permission.authorized_by_name or 'Líder',
                    'group_name': permission.group_name or 'Permiso Individual',
                    'user_name': user.get_full_name(),
                    'user_email': user.email,
                    'base_path': permission.base_path,
                    'total_users': 1,
                    'total_routes': 1,
                    'frontend_url': settings.FRONTEND_URL,
                }

                # Renderizar HTML para líder con template de MODIFICACIÓN
                html_content_leader = render_to_string('emails/permission_modified_leader.html', context_leader)

                # Preparar archivo adjunto
                timestamp = datetime.now().strftime('%Y%m%d')
                group_filename = f'{timestamp}_Permisos_Actualizados_{permission.group_name or "Individual"}.xlsx'

                # Enviar email al líder
                from django.core.mail import EmailMessage
                email_leader = EmailMessage(
                    subject=f'📧 Copia: Permisos Modificados - {permission.group_name or "Permiso Individual"} - Sistema IGAC',
                    body=html_content_leader,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[permission.authorized_by_email],
                )
                email_leader.content_subtype = 'html'
                email_leader.attach(group_filename, excel_file_leader.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                email_leader.send(fail_silently=False)

                emails_sent.append(permission.authorized_by_email)
                print(f"[OK] Email de actualización enviado al líder {permission.authorized_by_email}")

            except Exception as e:
                print(f"[ERROR] Error enviando email al líder {permission.authorized_by_email}: {str(e)}")
                import traceback
                traceback.print_exc()

        return Response({
            'message': 'Permiso actualizado exitosamente',
            'emails_sent': emails_sent,
            'permission': {
                'id': permission.id,
                'base_path': permission.base_path,
                'can_read': permission.can_read,
                'can_write': permission.can_write,
                'can_delete': permission.can_delete,
                'exempt_from_dictionary': permission.exempt_from_dictionary,
                'edit_permission_level': permission.edit_permission_level,
                'inheritance_mode': permission.inheritance_mode,
                'blocked_paths': permission.blocked_paths,
                'max_depth': permission.max_depth,
                'notes': permission.notes,
                'expires_at': permission.expires_at.isoformat() if permission.expires_at else None,
            }
        }, status=status.HTTP_200_OK)

    def delete_permission(self, request, permission_id=None):
        """
        DELETE /api/admin/permissions/{permission_id}/
        Elimina un permiso
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            permission = UserPermission.objects.get(pk=permission_id)
        except UserPermission.DoesNotExist:
            return Response(
                {'error': 'Permiso no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        base_path = permission.base_path
        permission.delete()

        return Response({
            'message': f'Permiso eliminado: {base_path}'
        })

    @action(detail=False, methods=['post'], url_path='permissions/(?P<permission_id>[^/.]+)/resend-email')
    def resend_permission_email(self, request, pk=None, permission_id=None):
        """
        POST /api/admin/permissions/{id}/resend-email
        Reenvía el correo de notificación de permisos al usuario
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            permission = UserPermission.objects.get(pk=permission_id, is_active=True)
        except UserPermission.DoesNotExist:
            return Response(
                {'error': 'Permiso no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        user = permission.user

        # Enviar email de notificación al usuario
        try:
            # Preparar contexto para el template
            base_path_escaped = permission.base_path.replace('/', '\\\\') if permission.base_path else ''
            context = {
                'user_name': user.get_full_name(),
                'base_path_display': f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if permission.base_path else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy",
                'can_read': permission.can_read,
                'can_write': permission.can_write,
                'can_delete': permission.can_delete,
                'can_create_directories': permission.can_create_directories,
                'exempt_from_dictionary': permission.exempt_from_dictionary,
                'edit_permission_level': permission.edit_permission_level,
                'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(permission.edit_permission_level, 'N/A') if permission.edit_permission_level else None,
                'inheritance_mode': permission.inheritance_mode,
                'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(permission.inheritance_mode),
                'granted_at': permission.granted_at.strftime('%d de %B de %Y a las %H:%M'),
                'expires_at': permission.expires_at.strftime('%d de %B de %Y'),
                'days_until_expiration': permission.days_until_expiration(),
                'granted_by': permission.granted_by.get_full_name() if permission.granted_by else 'Sistema',
                'granted_by_role': dict(User.ROLE_CHOICES).get(permission.granted_by.role) if permission.granted_by else 'Sistema',
                'frontend_url': settings.FRONTEND_URL,
                'notes': permission.notes,
            }

            # Renderizar HTML
            html_content = render_to_string('emails/permission_assigned.html', context)

            # Enviar email
            send_mail(
                subject=f'🔐 Permisos de Acceso Asignados - Sistema IGAC',
                message=f'Se le han asignado permisos de acceso. Vea los detalles en: {settings.FRONTEND_URL}',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_content,
                fail_silently=False,
            )
            print(f"[OK] Email de permiso reenviado exitosamente a {user.email}")

            return Response({
                'message': f'Correo enviado exitosamente a {user.email}'
            })

        except Exception as e:
            print(f"[ERROR] Error reenviando email de permiso a {user.email}: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al enviar correo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def download_permission_excel(self, request, permission_id=None):
        """
        GET /api/admin/permissions/{id}/download-excel/
        Descarga un reporte Excel de un permiso individual (una ruta)
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            permission = UserPermission.objects.select_related('user').get(pk=permission_id, is_active=True)
        except UserPermission.DoesNotExist:
            return Response(
                {'error': 'Permiso no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Generar Excel con un QuerySet (no una lista)
        from django.http import HttpResponse
        permissions_queryset = UserPermission.objects.filter(pk=permission_id, is_active=True).select_related('user')
        excel_file = self._generate_excel_report(permissions_queryset, permission.authorized_by_name or 'Sistema')

        # Preparar nombre de archivo
        timestamp = datetime.now().strftime('%Y%m%d')
        authorized_name = permission.authorized_by_name.replace(' ', '_') if permission.authorized_by_name else 'Sistema'
        filename = f'{timestamp}_Permisos_Repositorio_{authorized_name}.xlsx'

        # Crear respuesta HTTP
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['post'], url_path='bulk-assign-permissions')
    def bulk_assign_permissions(self, request):
        """
        POST /api/admin/bulk-assign-permissions
        Asigna permisos de forma masiva a múltiples usuarios y múltiples rutas

        Body: {
            "user_ids": [1, 2, 3, ...],
            "routes": [
                {
                    "base_path": "ruta1",
                    "blocked_paths": ["sub1", "sub2"],
                    "read_only_paths": ["sub3"]
                },
                {
                    "base_path": "ruta2",
                    "blocked_paths": [],
                    "read_only_paths": []
                }
            ],
            "can_read": true,
            "can_write": false,
            "can_delete": false,
            "can_create_directories": true,
            "exempt_from_dictionary": false,
            "edit_permission_level": "upload_only",
            "inheritance_mode": "total",
            "max_depth": null,
            "expires_at": "2025-12-31T23:59:59Z",
            "group_name": "Proyecto_X_2025",
            "notes": "..."
        }

        Returns: {
            "success": true,
            "total_assignments": 10,
            "users_notified": 5,
            "group_name": "Proyecto_X_2025",
            "assignments": [
                {
                    "user_id": 1,
                    "user_email": "user@igac.gov.co",
                    "routes": ["ruta1", "ruta2"],
                    "permission_ids": [101, 102]
                },
                ...
            ]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Validar datos con serializer
        serializer = BulkPermissionAssignmentSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {'error': 'Datos inválidos', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        data = serializer.validated_data

        # Extraer datos
        user_ids = data['user_ids']
        routes = data['routes']
        group_name = data['group_name']

        # Obtener usuarios
        users = User.objects.filter(id__in=user_ids, is_active=True)

        # Preparar fecha de vencimiento
        expires_at = data.get('expires_at')
        if not expires_at:
            current_year = timezone.now().year
            naive_dt = datetime(current_year, 12, 31, 23, 59, 59)
            expires_at = timezone.make_aware(naive_dt)

        # Variables para tracking
        total_assignments = 0
        users_notified = 0
        assignments = []
        failed_emails = []

        # Usar transacción para garantizar atomicidad
        try:
            with transaction.atomic():
                # Por cada usuario
                for user in users:
                    user_permissions = []
                    permission_ids = []

                    # Por cada ruta
                    for route_config in routes:
                        base_path_original = route_config['base_path']
                        blocked_paths = route_config.get('blocked_paths', [])
                        read_only_paths = route_config.get('read_only_paths', [])

                        print(f"\n[DEBUG] ===== NORMALIZANDO RUTA =====")
                        print(f"[DEBUG] Ruta ORIGINAL recibida: '{base_path_original}'")

                        # NORMALIZAR RUTA: Quitar prefijo completo si existe
                        # Primero, normalizar todas las barras a forward slash
                        base_path_normalized = base_path_original.replace('\\', '/')
                        print(f"[DEBUG] Ruta normalizada (/ en lugar de \\): '{base_path_normalized}'")

                        # Definir el prefijo base que debe removerse
                        prefix_to_remove = 'repositorio/DirGesCat/2510SP/H_Informacion_Consulta/Sub_Proy/'

                        # Remover el prefijo si existe (con o sin barras dobles al inicio)
                        if prefix_to_remove in base_path_normalized:
                            # Encontrar la posición del prefijo
                            idx = base_path_normalized.find(prefix_to_remove)
                            print(f"[DEBUG] Prefijo encontrado en posición: {idx}")
                            # Tomar todo después del prefijo
                            base_path = base_path_normalized[idx + len(prefix_to_remove):]
                            print(f"[DEBUG] Después de remover prefijo: '{base_path}'")
                        else:
                            # Si no se encontró el prefijo completo, usar la ruta normalizada
                            base_path = base_path_normalized
                            print(f"[DEBUG] NO se encontró el prefijo, usando ruta normalizada: '{base_path}'")

                        # Limpiar barras iniciales y finales
                        base_path = base_path.strip('/').strip('\\')
                        print(f"[DEBUG] Ruta FINAL a guardar en DB: '{base_path}'")
                        print(f"[DEBUG] =============================\n")

                        # Verificar si ya existe este permiso para el usuario (activo o inactivo)
                        existing_permission = UserPermission.objects.filter(
                            user=user,
                            base_path=base_path
                        ).first()

                        if existing_permission:
                            # Si existe (activo o inactivo), reactivarlo y actualizar
                            existing_permission.is_active = True
                            existing_permission.group_name = group_name
                            existing_permission.can_read = data.get('can_read', True)
                            existing_permission.can_write = data.get('can_write', False)
                            existing_permission.can_delete = data.get('can_delete', False)
                            existing_permission.can_create_directories = data.get('can_create_directories', True)
                            existing_permission.exempt_from_dictionary = data.get('exempt_from_dictionary', False)
                            existing_permission.edit_permission_level = data.get('edit_permission_level')
                            existing_permission.inheritance_mode = data.get('inheritance_mode', 'total')
                            existing_permission.blocked_paths = blocked_paths
                            existing_permission.read_only_paths = read_only_paths
                            existing_permission.max_depth = data.get('max_depth')
                            existing_permission.expires_at = expires_at
                            existing_permission.notes = data.get('notes', '')
                            existing_permission.authorized_by_email = data.get('authorized_by_email')
                            existing_permission.authorized_by_name = data.get('authorized_by_name')
                            existing_permission.granted_by = request.user
                            existing_permission.granted_at = timezone.now()
                            existing_permission.revoked_at = None
                            existing_permission.save()
                            permission = existing_permission
                            print(f"[INFO] Permiso reactivado para {user.email} en {base_path}")
                        else:
                            # Crear nuevo permiso solo si no existe ninguno
                            permission = UserPermission.objects.create(
                                user=user,
                                base_path=base_path,
                                group_name=group_name,
                                can_read=data.get('can_read', True),
                                can_write=data.get('can_write', False),
                                can_delete=data.get('can_delete', False),
                                can_create_directories=data.get('can_create_directories', True),
                                exempt_from_dictionary=data.get('exempt_from_dictionary', False),
                                edit_permission_level=data.get('edit_permission_level'),
                                inheritance_mode=data.get('inheritance_mode', 'total'),
                                blocked_paths=blocked_paths,
                                read_only_paths=read_only_paths,
                                max_depth=data.get('max_depth'),
                                expires_at=expires_at,
                                granted_by=request.user,
                                notes=data.get('notes', ''),
                                authorized_by_email=data.get('authorized_by_email'),
                                authorized_by_name=data.get('authorized_by_name')
                            )
                            print(f"[INFO] Permiso creado para {user.email} en {base_path}")

                        user_permissions.append(permission)
                        permission_ids.append(permission.id)
                        total_assignments += 1

                    # Enviar correo individual a cada usuario informando sus permisos
                    try:
                        # Preparar lista de rutas para el template
                        routes_for_template = []
                        for perm in user_permissions:
                            base_path_escaped = perm.base_path.replace('/', '\\\\') if perm.base_path else ''
                            route_data = {
                                'base_path_display': f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if perm.base_path else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy",
                                'blocked_paths': perm.blocked_paths,
                                'read_only_paths': perm.read_only_paths
                            }
                            routes_for_template.append(route_data)

                        # Usar el primer permiso para datos generales (todos tienen los mismos permisos)
                        first_perm = user_permissions[0]

                        # Preparar contexto para el template
                        context = {
                            'user_name': user.get_full_name(),
                            'group_name': group_name,
                            'routes': routes_for_template,
                            'can_read': first_perm.can_read,
                            'can_write': first_perm.can_write,
                            'can_delete': first_perm.can_delete,
                            'can_create_directories': first_perm.can_create_directories,
                            'exempt_from_dictionary': first_perm.exempt_from_dictionary,
                            'edit_permission_level': first_perm.edit_permission_level,
                            'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(first_perm.edit_permission_level, 'N/A') if first_perm.edit_permission_level else None,
                            'inheritance_mode': first_perm.inheritance_mode,
                            'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(first_perm.inheritance_mode),
                            'granted_at': first_perm.granted_at.strftime('%d de %B de %Y a las %H:%M'),
                            'expires_at': first_perm.expires_at.strftime('%d de %B de %Y'),
                            'days_until_expiration': first_perm.days_until_expiration(),
                            'granted_by': request.user.get_full_name(),
                            'granted_by_role': dict(User.ROLE_CHOICES).get(request.user.role),
                            'frontend_url': settings.FRONTEND_URL,
                            'notes': first_perm.notes,
                        }

                        # Renderizar HTML con template de bulk
                        html_content = render_to_string('emails/permission_bulk_assigned.html', context)

                        # Generar Excel individual con los permisos del usuario
                        excel_file = self._generate_excel_report(user_permissions, first_perm.authorized_by_name or 'Sistema')

                        # Preparar nombre de archivo
                        timestamp = datetime.now().strftime('%Y%m%d')
                        authorized_name = first_perm.authorized_by_name.replace(' ', '_') if first_perm.authorized_by_name else 'Sistema'
                        safe_user_name = user.get_full_name().replace(' ', '_')
                        filename = f'{timestamp}_Permisos_{safe_user_name}_{authorized_name}.xlsx'

                        # Crear email con adjunto
                        from django.core.mail import EmailMessage
                        email = EmailMessage(
                            subject=f'🔐 Permisos de Acceso Asignados ({len(routes)} rutas) - Sistema IGAC',
                            body=html_content,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[user.email],
                        )
                        email.content_subtype = 'html'

                        # Adjuntar Excel individual
                        email.attach(filename, excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                        # Enviar correo individual
                        email.send(fail_silently=False)
                        users_notified += 1
                        print(f"[OK] Email individual enviado a {user.email} ({len(routes)} rutas)")
                    except Exception as e:
                        # Log el error pero no fallar la operación
                        failed_emails.append(user.email)
                        print(f"[ERROR] Error enviando email individual a {user.email}: {str(e)}")
                        import traceback
                        traceback.print_exc()

                    # Agregar a resultados
                    assignments.append({
                        'user_id': user.id,
                        'user_email': user.email,
                        'user_name': user.get_full_name(),
                        'routes': [r['base_path'] for r in routes],
                        'permission_ids': permission_ids
                    })

                # ==================== ENVIAR UN SOLO CORREO AL SOLICITANTE ====================
                # Enviar correo SOLO al solicitante (authorized_by_email) con toda la información del grupo
                authorized_email = data.get('authorized_by_email')
                authorized_name = data.get('authorized_by_name') or 'Solicitante'

                if authorized_email:
                    try:
                        # Obtener todos los permisos del grupo
                        all_group_permissions = UserPermission.objects.filter(
                            group_name=group_name,
                            is_active=True
                        ).select_related('user').order_by('user__email', 'base_path')

                        # Generar Excel del grupo completo
                        excel_file = self._generate_group_excel_report_by_route(
                            all_group_permissions,
                            authorized_name,
                            group_name
                        )

                        # Preparar nombre de archivo
                        timestamp = datetime.now().strftime('%Y%m%d')
                        safe_authorized_name = authorized_name.replace(' ', '_')
                        safe_group_name = group_name.replace(' ', '_')
                        filename = f'{timestamp}_Permisos_Grupo_{safe_group_name}_{safe_authorized_name}.xlsx'

                        # Preparar lista de usuarios para el correo
                        users_list = [f"{a['user_name']} ({a['user_email']})" for a in assignments]
                        routes_list = [r['base_path'] for r in routes]

                        # Crear email simple con información del grupo
                        from django.core.mail import EmailMessage
                        
                        # Construir lista de rutas con backslashes (NO dentro de f-string)
                        routes_html = "".join([f"<li>\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{r}</li>" for r in routes_list])
                        
                        email_body = f"""
                        <html>
                        <body style="font-family: Arial, sans-serif;">
                            <h2>Asignación Masiva de Permisos - Grupo: {group_name}</h2>

                            <h3>Resumen:</h3>
                            <ul>
                                <li><strong>Total de usuarios:</strong> {len(users)}</li>
                                <li><strong>Total de rutas:</strong> {len(routes)}</li>
                                <li><strong>Permisos asignados:</strong> {total_assignments}</li>
                            </ul>

                            <h3>Usuarios con acceso:</h3>
                            <ul>
                                {"".join([f"<li>{u}</li>" for u in users_list])}
                            </ul>

                            <h3>Rutas autorizadas:</h3>
                            <ul>
                                {routes_html}
                            </ul>

                            <p><strong>Adjunto:</strong> Excel con el detalle completo de permisos.</p>

                            <hr>
                            <p style="color: #666; font-size: 12px;">
                                Sistema de Gestión de Archivos NetApp - IGAC<br>
                                Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}
                            </p>
                        </body>
                        </html>
                        """

                        email = EmailMessage(
                            subject=f'📋 Reporte de Asignación Masiva - Grupo "{group_name}" - {len(users)} usuarios - Sistema IGAC',
                            body=email_body,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[authorized_email],
                        )
                        email.content_subtype = 'html'

                        # Adjuntar Excel del grupo
                        email.attach(filename, excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                        # Enviar
                        email.send(fail_silently=False)
                        users_notified = 1  # Solo el solicitante
                        print(f"[OK] Email con Excel del grupo enviado exitosamente a {authorized_email}")
                        print(f"     Grupo: {group_name} - {len(users)} usuarios - {len(routes)} rutas - {total_assignments} permisos")
                    except Exception as e:
                        # Log el error pero no fallar la operación
                        failed_emails.append(authorized_email)
                        print(f"[ERROR] Error enviando email al solicitante {authorized_email}: {str(e)}")
                        import traceback
                        traceback.print_exc()
                else:
                    print(f"[WARNING] No se proporcionó email del solicitante, no se envió correo de notificación")

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',  # Usar acción existente (max 20 chars)
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Grupo: {group_name}',
                    details={
                        'action_type': 'bulk_assign_permissions',
                        'group_name': group_name,
                        'user_count': len(users),
                        'route_count': len(routes),
                        'total_assignments': total_assignments,
                        'users_notified': users_notified,
                        'description': f'Asignación masiva: Grupo "{group_name}" - {len(users)} usuarios - {len(routes)} rutas - {total_assignments} permisos'
                    },
                    success=True
                )

        except Exception as e:
            # Si algo falla, hacer rollback
            print(f"[ERROR] Error en asignación masiva: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error en la asignación masiva: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Respuesta exitosa
        response_data = {
            'success': True,
            'total_assignments': total_assignments,
            'users_notified': users_notified,
            'group_name': group_name,
            'assignments': assignments
        }

        if failed_emails:
            response_data['warning'] = f'Se crearon los permisos pero falló el envío de emails a: {", ".join(failed_emails)}'

        return Response(response_data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='groups')
    def list_groups(self, request):
        """
        GET /api/admin/groups/
        Lista todos los grupos con estadísticas

        Returns: {
            "groups": [
                {
                    "group_name": "Proyecto_X_2025",
                    "user_count": 5,
                    "route_count": 2,
                    "total_permissions": 10,
                    "created_at": "2025-01-15T10:30:00Z",
                    "granted_by": "admin@igac.gov.co",
                    "expires_at": "2025-12-31T23:59:59Z"
                },
                ...
            ]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener grupos únicos
        from django.db.models import Count, Min, Max
        groups = UserPermission.objects.filter(
            group_name__isnull=False,
            is_active=True
        ).values('group_name').annotate(
            user_count=Count('user', distinct=True),
            route_count=Count('base_path', distinct=True),
            total_permissions=Count('id'),
            created_at=Min('granted_at'),
            expires_at=Max('expires_at')
        ).order_by('-created_at')

        # Enriquecer con información adicional
        groups_data = []
        for group in groups:
            # Obtener el granted_by más reciente
            recent_permission = UserPermission.objects.filter(
                group_name=group['group_name'],
                is_active=True
            ).select_related('granted_by').order_by('-granted_at').first()

            # Obtener emails únicos de usuarios en el grupo (para búsqueda)
            user_emails = list(
                UserPermission.objects.filter(
                    group_name=group['group_name'],
                    is_active=True
                ).select_related('user').values_list('user__email', flat=True).distinct()
            )

            groups_data.append({
                'group_name': group['group_name'],
                'user_count': group['user_count'],
                'route_count': group['route_count'],
                'total_permissions': group['total_permissions'],
                'created_at': group['created_at'],
                'granted_by': recent_permission.granted_by.get_full_name() if recent_permission and recent_permission.granted_by else 'Sistema',
                'expires_at': group['expires_at'],
                'user_emails': user_emails  # Lista de emails para búsqueda
            })

        return Response({'groups': groups_data})

    @action(detail=False, methods=['get'], url_path='groups/(?P<group_name>[^/.]+)/permissions')
    def get_group_permissions(self, request, group_name=None):
        """
        GET /api/admin/groups/{group_name}/permissions/
        Obtiene todos los permisos de un grupo específico

        Returns: {
            "group_name": "Proyecto_X_2025",
            "permissions": [
                {
                    "id": 101,
                    "user": { ... },
                    "base_path": "ruta1",
                    "can_read": true,
                    ...
                },
                ...
            ],
            "users": [...],
            "routes": [...]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener permisos del grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        ).select_related('user', 'granted_by').order_by('user__email', 'base_path')

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Serializar permisos
        serializer = UserPermissionSerializer(permissions, many=True)

        # Obtener usuarios únicos
        users = list(set([perm.user for perm in permissions]))
        users_data = [{'id': u.id, 'email': u.email, 'name': u.get_full_name()} for u in users]

        # Obtener rutas únicas
        routes = list(set([perm.base_path for perm in permissions]))

        return Response({
            'group_name': group_name,
            'permissions': serializer.data,
            'users': users_data,
            'routes': routes
        })

    @action(detail=False, methods=['patch'], url_path='groups/(?P<group_name>[^/.]+)/update')
    def update_group_permissions(self, request, group_name=None):
        """
        PATCH /api/admin/groups/{group_name}/update/
        Actualiza los permisos de todos los usuarios en un grupo

        Body: {
            "can_read": true,
            "can_write": false,
            ...
            "expires_at": "2025-12-31T23:59:59Z"
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener permisos del grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data
        updated_count = 0

        try:
            with transaction.atomic():
                # Actualizar todos los permisos del grupo
                for permission in permissions:
                    if 'can_read' in data:
                        permission.can_read = data['can_read']
                    if 'can_write' in data:
                        permission.can_write = data['can_write']
                    if 'can_delete' in data:
                        permission.can_delete = data['can_delete']
                    if 'can_create_directories' in data:
                        permission.can_create_directories = data['can_create_directories']
                    if 'exempt_from_dictionary' in data:
                        permission.exempt_from_dictionary = data['exempt_from_dictionary']
                    if 'edit_permission_level' in data:
                        permission.edit_permission_level = data['edit_permission_level']
                    if 'inheritance_mode' in data:
                        permission.inheritance_mode = data['inheritance_mode']
                    if 'max_depth' in data:
                        permission.max_depth = data['max_depth']
                    if 'expires_at' in data:
                        permission.expires_at = data['expires_at']
                    if 'notes' in data:
                        permission.notes = data['notes']
                    if 'authorized_by_email' in data:
                        permission.authorized_by_email = data['authorized_by_email']
                    if 'authorized_by_name' in data:
                        permission.authorized_by_name = data['authorized_by_name']

                    permission.save()
                    updated_count += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Actualizar permisos del grupo: {group_name}',
                    details={
                        'action_type': 'update_group_permissions',
                        'group_name': group_name,
                        'updated_count': updated_count,
                        'description': f'Actualización de permisos del grupo "{group_name}" - {updated_count} permisos actualizados'
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al actualizar permisos: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'updated_count': updated_count,
            'message': f'Se actualizaron {updated_count} permisos del grupo "{group_name}"'
        })

    @action(detail=False, methods=['delete'], url_path='groups/(?P<group_name>[^/.]+)')
    def delete_group(self, request, group_name=None):
        """
        DELETE /api/admin/groups/{group_name}/
        Elimina (desactiva) todos los permisos de un grupo
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener permisos del grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        deleted_count = 0

        try:
            with transaction.atomic():
                from django.utils import timezone
                now = timezone.now()

                # Desactivar todos los permisos
                for permission in permissions:
                    permission.is_active = False
                    permission.revoked_at = now
                    permission.save()
                    deleted_count += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='delete',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Eliminar grupo: {group_name}',
                    details={
                        'action_type': 'delete_group',
                        'group_name': group_name,
                        'deleted_count': deleted_count,
                        'description': f'Eliminación del grupo "{group_name}" - {deleted_count} permisos desactivados'
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al eliminar grupo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'deleted_count': deleted_count,
            'message': f'Se eliminaron {deleted_count} permisos del grupo "{group_name}"'
        })

    def download_group_excel(self, request, group_name=None):
        """
        GET /api/admin/groups/{group_name}/download-excel/
        Descarga un reporte Excel de todos los permisos de un grupo
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener permisos del grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        ).select_related('user').order_by('user__email', 'base_path')

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Tomar el authorized_by_name del primer permiso (todos deberían tener el mismo)
        first_perm = permissions.first()
        authorized_name = first_perm.authorized_by_name or 'Sistema'

        # Generar Excel AGRUPADO POR RUTA (no por usuario)
        from django.http import HttpResponse
        excel_file = self._generate_group_excel_report_by_route(permissions, authorized_name, group_name)

        # Preparar nombre de archivo
        timestamp = datetime.now().strftime('%Y%m%d')
        authorized_name_clean = authorized_name.replace(' ', '_')
        filename = f'{timestamp}_Permisos_Repositorio_{authorized_name_clean}.xlsx'

        # Crear respuesta HTTP
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'], url_path='authorization-autocomplete')
    def get_authorization_autocomplete(self, request):
        """
        GET /api/admin/users/authorization-autocomplete
        Retorna emails y nombres únicos de personas que han autorizado permisos
        Para autocomplete en formularios

        Returns: {
            "authorizers": [
                {
                    "email": "leader@igac.gov.co",
                    "name": "Juan Pérez"
                },
                ...
            ]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener combinaciones únicas de email y nombre
        authorizers = UserPermission.objects.filter(
            authorized_by_email__isnull=False,
            authorized_by_name__isnull=False
        ).values('authorized_by_email', 'authorized_by_name').distinct().order_by('authorized_by_name')

        # Formatear respuesta
        authorizers_list = [
            {
                'email': auth['authorized_by_email'],
                'name': auth['authorized_by_name']
            }
            for auth in authorizers
        ]

        return Response({'authorizers': authorizers_list})

    @action(detail=False, methods=['delete'], url_path='groups/(?P<group_name>[^/.]+)/remove-user/(?P<user_id>[^/.]+)')
    def remove_user_from_group(self, request, group_name=None, user_id=None):
        """
        DELETE /api/admin/groups/{group_name}/remove-user/{user_id}/
        Elimina un usuario específico de un grupo
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Obtener permisos del usuario en el grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            user_id=user_id,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para el usuario en el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        deleted_count = 0

        try:
            with transaction.atomic():
                from django.utils import timezone
                now = timezone.now()

                # Desactivar permisos
                for permission in permissions:
                    permission.is_active = False
                    permission.revoked_at = now
                    permission.save()
                    deleted_count += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='delete',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Remover usuario del grupo: {group_name}',
                    details={
                        'action_type': 'remove_user_from_group',
                        'group_name': group_name,
                        'user_id': user_id,
                        'deleted_count': deleted_count,
                        'description': f'Usuario {user_id} removido del grupo "{group_name}" - {deleted_count} permisos desactivados'
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al remover usuario: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'user_id': int(user_id),
            'deleted_count': deleted_count
        })

    @action(detail=False, methods=['post'], url_path='groups/(?P<group_name>[^/.]+)/add-users')
    def add_users_to_group(self, request, group_name=None):
        """
        POST /api/admin/groups/{group_name}/add-users/
        Agrega nuevos usuarios a un grupo existente
        Les asigna TODAS las rutas del grupo con los mismos permisos

        Body: {
            "user_ids": [1, 2, 3]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        user_ids = request.data.get('user_ids', [])

        if not user_ids:
            return Response(
                {'error': 'Debe proporcionar al menos un user_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar que el grupo existe
        existing_permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        ).select_related('user', 'granted_by')

        if not existing_permissions.exists():
            return Response(
                {'error': f'No se encontró el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener rutas únicas del grupo con sus permisos
        routes_with_permissions = {}
        for perm in existing_permissions:
            if perm.base_path not in routes_with_permissions:
                # Guardar los permisos de esta ruta (tomar los de cualquier usuario, son los mismos)
                routes_with_permissions[perm.base_path] = {
                    'can_read': perm.can_read,
                    'can_write': perm.can_write,
                    'can_delete': perm.can_delete,
                    'can_create_directories': perm.can_create_directories,
                    'exempt_from_dictionary': perm.exempt_from_dictionary,
                    'edit_permission_level': perm.edit_permission_level,
                    'inheritance_mode': perm.inheritance_mode,
                    'max_depth': perm.max_depth,
                    'expires_at': perm.expires_at,
                    'notes': perm.notes,
                    'blocked_paths': perm.blocked_paths,
                    'read_only_paths': perm.read_only_paths,
                }

        # Obtener usuarios a agregar
        users = User.objects.filter(id__in=user_ids)

        if users.count() != len(user_ids):
            return Response(
                {'error': 'Algunos usuarios no fueron encontrados'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Filtrar usuarios que YA están en el grupo
        existing_user_ids = set(existing_permissions.values_list('user_id', flat=True).distinct())
        new_users = [u for u in users if u.id not in existing_user_ids]

        if not new_users:
            return Response(
                {'error': 'Todos los usuarios ya pertenecen al grupo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        total_assignments = 0
        users_notified = 0

        try:
            with transaction.atomic():
                # Crear/reactivar permisos para cada nuevo usuario con cada ruta
                for user in new_users:
                    user_routes = []

                    for base_path, perm_config in routes_with_permissions.items():
                        # Buscar si ya existe un permiso (activo o inactivo) para este user+path
                        # NO filtrar por group_name porque puede existir de otra asignación
                        existing_permission = UserPermission.objects.filter(
                            user=user,
                            base_path=base_path
                        ).first()

                        if existing_permission:
                            # Si existe, actualizar/reactivar el permiso existente
                            existing_permission.is_active = True
                            existing_permission.revoked_at = None
                            existing_permission.granted_by = request.user
                            existing_permission.group_name = group_name  # Actualizar al grupo actual
                            # Actualizar con los permisos actuales del grupo
                            existing_permission.can_read = perm_config['can_read']
                            existing_permission.can_write = perm_config['can_write']
                            existing_permission.can_delete = perm_config['can_delete']
                            existing_permission.can_create_directories = perm_config['can_create_directories']
                            existing_permission.exempt_from_dictionary = perm_config['exempt_from_dictionary']
                            existing_permission.edit_permission_level = perm_config['edit_permission_level']
                            existing_permission.inheritance_mode = perm_config['inheritance_mode']
                            existing_permission.max_depth = perm_config['max_depth']
                            existing_permission.expires_at = perm_config['expires_at']
                            existing_permission.notes = perm_config['notes']
                            existing_permission.blocked_paths = perm_config['blocked_paths']
                            existing_permission.read_only_paths = perm_config['read_only_paths']
                            existing_permission.save()
                            permission = existing_permission
                        else:
                            # CREAR nuevo permiso solo si no existe ninguno
                            permission = UserPermission.objects.create(
                                user=user,
                                base_path=base_path,
                                granted_by=request.user,
                                group_name=group_name,
                                **perm_config
                            )

                        total_assignments += 1
                        user_routes.append({
                            'base_path': base_path,
                            'blocked_paths': perm_config['blocked_paths'],
                            'read_only_paths': perm_config['read_only_paths'],
                        })

                    # Enviar email SOLO a este usuario nuevo
                    try:
                        from django.core.mail import send_mail
                        from django.template.loader import render_to_string
                        from django.conf import settings

                        # Preparar lista de rutas para el template
                        routes_for_template = []
                        for route_data in user_routes:
                            base_path_escaped = route_data['base_path'].replace('/', '\\\\') if route_data['base_path'] else ''
                            route_display = {
                                'base_path_display': f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_escaped}" if route_data['base_path'] else "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy",
                                'blocked_paths': route_data.get('blocked_paths', []),
                                'read_only_paths': route_data.get('read_only_paths', [])
                            }
                            routes_for_template.append(route_display)

                        # Obtener el primer permiso creado para datos generales
                        first_perm = UserPermission.objects.filter(
                            user=user,
                            group_name=group_name,
                            is_active=True
                        ).first()

                        if first_perm:
                            # Preparar contexto para el template
                            context = {
                                'user_name': user.get_full_name(),
                                'group_name': group_name,
                                'routes': routes_for_template,
                                'can_read': first_perm.can_read,
                                'can_write': first_perm.can_write,
                                'can_delete': first_perm.can_delete,
                                'can_create_directories': first_perm.can_create_directories,
                                'exempt_from_dictionary': first_perm.exempt_from_dictionary,
                                'edit_permission_level': first_perm.edit_permission_level,
                                'edit_permission_level_display': dict(UserPermission.EDIT_PERMISSION_CHOICES).get(first_perm.edit_permission_level, 'N/A') if first_perm.edit_permission_level else None,
                                'inheritance_mode': first_perm.inheritance_mode,
                                'inheritance_mode_display': dict(UserPermission.INHERITANCE_MODE_CHOICES).get(first_perm.inheritance_mode),
                                'granted_at': first_perm.granted_at.strftime('%d de %B de %Y a las %H:%M'),
                                'expires_at': first_perm.expires_at.strftime('%d de %B de %Y'),
                                'days_until_expiration': first_perm.days_until_expiration(),
                                'granted_by': request.user.get_full_name(),
                                'granted_by_role': dict(User.ROLE_CHOICES).get(request.user.role),
                                'frontend_url': settings.FRONTEND_URL,
                                'notes': first_perm.notes,
                            }

                            # Renderizar HTML con template de bulk
                            html_content = render_to_string('emails/permission_bulk_assigned.html', context)

                            # Enviar email
                            send_mail(
                                subject=f'🔐 Permisos de Acceso Asignados ({len(user_routes)} rutas) - Sistema IGAC',
                                message=f'Se le han asignado permisos de acceso a {len(user_routes)} rutas. Vea los detalles en: {settings.FRONTEND_URL}',
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[user.email],
                                html_message=html_content,
                                fail_silently=False,
                            )
                            users_notified += 1
                            print(f"[OK] Email enviado exitosamente a {user.email} ({len(user_routes)} rutas)")
                    except Exception as e:
                        print(f"[ERROR] Error enviando email a {user.email}: {str(e)}")
                        import traceback
                        traceback.print_exc()

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Agregar usuarios al grupo: {group_name}',
                    details={
                        'action_type': 'add_users_to_group',
                        'group_name': group_name,
                        'new_users_count': len(new_users),
                        'total_assignments': total_assignments,
                        'users_notified': users_notified,
                        'user_ids': [u.id for u in new_users]
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al agregar usuarios: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'new_users_count': len(new_users),
            'total_assignments': total_assignments,
            'users_notified': users_notified,
            'message': f'Se agregaron {len(new_users)} usuario(s) al grupo con {len(routes_with_permissions)} ruta(s)'
        })

    @action(detail=False, methods=['post'], url_path='groups/(?P<group_name>[^/.]+)/add-routes')
    def add_routes_to_group(self, request, group_name=None):
        """
        POST /api/admin/groups/{group_name}/add-routes/
        Agrega nuevas rutas a un grupo existente
        Las asigna a TODOS los usuarios del grupo con los mismos permisos que las rutas existentes

        Body: {
            "routes": [
                {
                    "base_path": "nueva/ruta",
                    "blocked_paths": [],
                    "read_only_paths": []
                }
            ]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        routes = request.data.get('routes', [])

        if not routes:
            return Response(
                {'error': 'Debe proporcionar al menos una ruta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar que el grupo existe
        existing_permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        ).select_related('user')

        if not existing_permissions.exists():
            return Response(
                {'error': f'No se encontró el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener permisos del grupo (tomar los de cualquier permiso, son los mismos)
        sample_perm = existing_permissions.first()
        group_permissions = {
            'can_read': sample_perm.can_read,
            'can_write': sample_perm.can_write,
            'can_delete': sample_perm.can_delete,
            'can_create_directories': sample_perm.can_create_directories,
            'exempt_from_dictionary': sample_perm.exempt_from_dictionary,
            'edit_permission_level': sample_perm.edit_permission_level,
            'inheritance_mode': sample_perm.inheritance_mode,
            'max_depth': sample_perm.max_depth,
            'expires_at': sample_perm.expires_at,
            'notes': sample_perm.notes,
        }

        # Obtener usuarios únicos del grupo
        users = list(set([perm.user for perm in existing_permissions]))

        # Obtener rutas existentes del grupo
        existing_routes = set(existing_permissions.values_list('base_path', flat=True).distinct())

        # NORMALIZAR RUTAS: Quitar prefijo completo si existe
        normalized_routes = []
        for route in routes:
            base_path_original = route['base_path']

            print(f"\n[DEBUG] ===== NORMALIZANDO RUTA (add_routes_to_group) =====")
            print(f"[DEBUG] Ruta ORIGINAL recibida: '{base_path_original}'")

            # Normalizar todas las barras a forward slash
            base_path_normalized = base_path_original.replace('\\', '/')
            print(f"[DEBUG] Ruta normalizada (/ en lugar de \\): '{base_path_normalized}'")

            # Definir el prefijo base que debe removerse
            prefix_to_remove = 'repositorio/DirGesCat/2510SP/H_Informacion_Consulta/Sub_Proy/'

            # Remover el prefijo si existe
            if prefix_to_remove in base_path_normalized:
                idx = base_path_normalized.find(prefix_to_remove)
                print(f"[DEBUG] Prefijo encontrado en posición: {idx}")
                base_path = base_path_normalized[idx + len(prefix_to_remove):]
                print(f"[DEBUG] Después de remover prefijo: '{base_path}'")
            else:
                base_path = base_path_normalized
                print(f"[DEBUG] NO se encontró el prefijo, usando ruta normalizada: '{base_path}'")

            # Limpiar barras iniciales y finales
            base_path = base_path.strip('/').strip('\\')
            print(f"[DEBUG] Ruta FINAL a guardar en DB: '{base_path}'")
            print(f"[DEBUG] =============================\n")

            normalized_routes.append({
                'base_path': base_path,
                'blocked_paths': route.get('blocked_paths', []),
                'read_only_paths': route.get('read_only_paths', [])
            })

        # Filtrar rutas nuevas (que no existen en el grupo)
        new_routes = [r for r in normalized_routes if r['base_path'] not in existing_routes]

        if not new_routes:
            return Response(
                {'error': 'Todas las rutas ya existen en el grupo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        total_assignments = 0

        try:
            with transaction.atomic():
                # Crear permisos para cada usuario con cada ruta nueva
                for user in users:
                    for route in new_routes:
                        UserPermission.objects.create(
                            user=user,
                            base_path=route['base_path'],
                            granted_by=request.user,
                            group_name=group_name,
                            blocked_paths=route.get('blocked_paths', []),
                            read_only_paths=route.get('read_only_paths', []),
                            **group_permissions
                        )
                        total_assignments += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Agregar rutas al grupo: {group_name}',
                    details={
                        'action_type': 'add_routes_to_group',
                        'group_name': group_name,
                        'new_routes_count': len(new_routes),
                        'total_assignments': total_assignments,
                        'users_affected': len(users),
                        'new_routes': [r['base_path'] for r in new_routes]
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al agregar rutas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # ENVIAR NOTIFICACIONES POR EMAIL (si notify_users=True)
        notify_users_flag = request.data.get('notify_users', False)
        notify_leader_flag = request.data.get('notify_leader', False)
        emails_sent = []

        if notify_users_flag:
            print(f"\n[EMAIL] Enviando notificaciones a {len(users)} usuario(s) del grupo {group_name}")

            for user in users:
                try:
                    # Preparar rutas para mostrar en el email
                    routes_display = ["\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\" + r['base_path'].replace('/', '\\\\') for r in new_routes]

                    context = {
                        'user_name': user.get_full_name(),
                        'group_name': group_name,
                        'new_routes': routes_display,
                        'frontend_url': settings.FRONTEND_URL,
                    }

                    # Renderizar HTML
                    html_content = render_to_string('emails/group_routes_added.html', context)

                    # Enviar email
                    from django.core.mail import EmailMessage
                    email = EmailMessage(
                        subject=f'✨ Nuevas Rutas Agregadas - Grupo {group_name} - Sistema IGAC',
                        body=html_content,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[user.email],
                    )
                    email.content_subtype = 'html'
                    email.send(fail_silently=False)

                    emails_sent.append(user.email)
                    print(f"[OK] Email enviado a {user.email}")

                except Exception as e:
                    print(f"[ERROR] Error enviando email a {user.email}: {str(e)}")
                    import traceback
                    traceback.print_exc()

        # ENVIAR EMAIL AL LÍDER (si notify_leader=True y existe authorized_by_email en el grupo)
        if notify_leader_flag and existing_permissions.exists():
            # Obtener el authorized_by_email del primer permiso (todos deberían tener el mismo)
            sample_perm = existing_permissions.first()
            if sample_perm.authorized_by_email:
                try:
                    routes_display = ["\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\" + r['base_path'].replace('/', '\\\\') for r in new_routes]

                    context_leader = {
                        'leader_name': sample_perm.authorized_by_name or 'Líder',
                        'group_name': group_name,
                        'new_routes': routes_display,
                        'new_routes_count': len(new_routes),
                        'users_affected': len(users),
                        'frontend_url': settings.FRONTEND_URL,
                    }

                    html_content_leader = render_to_string('emails/group_routes_added.html', context_leader)

                    from django.core.mail import EmailMessage
                    email_leader = EmailMessage(
                        subject=f'📧 Copia: Nuevas Rutas en Grupo {group_name} - Sistema IGAC',
                        body=html_content_leader,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[sample_perm.authorized_by_email],
                    )
                    email_leader.content_subtype = 'html'
                    email_leader.send(fail_silently=False)

                    emails_sent.append(sample_perm.authorized_by_email)
                    print(f"[OK] Email enviado al líder {sample_perm.authorized_by_email}")

                except Exception as e:
                    print(f"[ERROR] Error enviando email al líder: {str(e)}")
                    import traceback
                    traceback.print_exc()

        return Response({
            'success': True,
            'group_name': group_name,
            'new_routes_count': len(new_routes),
            'total_assignments': total_assignments,
            'users_affected': len(users),
            'emails_sent': emails_sent,
            'message': f'Se agregaron {len(new_routes)} ruta(s) al grupo para {len(users)} usuario(s)'
        })

    def update_route_permissions(self, request, group_name=None, route_path=None):
        """
        PATCH /api/admin/groups/{group_name}/routes/{route_path}/permissions/
        Actualiza los permisos de una ruta específica para TODOS los usuarios del grupo

        Body: {
            "can_read": true,
            "can_write": false,
            ...
        }
        """
        print(f"[DEBUG] update_route_permissions called!")
        print(f"[DEBUG] group_name: {group_name}")
        print(f"[DEBUG] route_path (raw): {route_path}")
        print(f"[DEBUG] request.method: {request.method}")

        error = self._check_superadmin(request.user)
        if error:
            return error

        # Decodificar la ruta (viene en URL encoding)
        from urllib.parse import unquote
        route_path = unquote(route_path)

        # Normalizar ruta para display en emails (reemplazar / por \)
        route_path_normalized = route_path.replace('/', '\\')

        # Obtener permisos de esta ruta en el grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            base_path=route_path,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para la ruta "{route_path}" en el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data
        updated_count = 0

        try:
            with transaction.atomic():
                # Actualizar permisos de TODOS los usuarios para esta ruta
                for permission in permissions:
                    if 'can_read' in data:
                        permission.can_read = data['can_read']
                    if 'can_write' in data:
                        permission.can_write = data['can_write']
                    if 'can_delete' in data:
                        permission.can_delete = data['can_delete']
                    if 'can_create_directories' in data:
                        permission.can_create_directories = data['can_create_directories']
                    if 'exempt_from_dictionary' in data:
                        permission.exempt_from_dictionary = data['exempt_from_dictionary']
                    if 'edit_permission_level' in data:
                        permission.edit_permission_level = data['edit_permission_level']
                    if 'inheritance_mode' in data:
                        permission.inheritance_mode = data['inheritance_mode']
                    if 'max_depth' in data:
                        permission.max_depth = data['max_depth']
                    if 'blocked_paths' in data:
                        permission.blocked_paths = data['blocked_paths']
                    if 'read_only_paths' in data:
                        permission.read_only_paths = data['read_only_paths']
                    if 'authorized_by_email' in data:
                        permission.authorized_by_email = data['authorized_by_email']
                    if 'authorized_by_name' in data:
                        permission.authorized_by_name = data['authorized_by_name']

                    permission.save()
                    updated_count += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',
                    target_path=f'bulk_permissions/{group_name}/{route_path}',
                    target_name=f'Modificar permisos de ruta en grupo: {group_name}',
                    details={
                        'action_type': 'update_route_permissions',
                        'group_name': group_name,
                        'route_path': route_path,
                        'updated_count': updated_count,
                        'changes': data
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al actualizar permisos: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # ENVIAR NOTIFICACIONES POR EMAIL
        notify_users_flag = data.get('notify_users', False)
        notify_leader_flag = data.get('notify_leader', False)
        emails_sent = []

        # Generar Excel del grupo completo (se usará para todos los emails)
        excel_file = None
        if notify_users_flag or notify_leader_flag:
            try:
                all_group_permissions = UserPermission.objects.filter(
                    group_name=group_name,
                    is_active=True
                ).select_related('user').order_by('user__email', 'base_path')

                sample_perm = permissions.first()
                authorized_by_name = sample_perm.authorized_by_name if sample_perm else 'Sistema'

                excel_file = self._generate_group_excel_report_by_route(
                    all_group_permissions,
                    authorized_by_name,
                    group_name
                )
                print(f"[OK] Excel del grupo generado para notificaciones")
            except Exception as e:
                print(f"[ERROR] Error generando Excel del grupo: {str(e)}")

        if notify_users_flag:
            print(f"\n[EMAIL] Enviando notificaciones de modificación de permisos a usuarios")

            # Obtener los permisos actualizados
            for permission in permissions:
                try:
                    route_display = f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{route_path_normalized}"

                    context = {
                        'user_name': permission.user.get_full_name(),
                        'group_name': group_name,
                        'route_display': route_display,
                        'can_read': permission.can_read,
                        'can_write': permission.can_write,
                        'can_delete': permission.can_delete,
                        'can_create_directories': permission.can_create_directories,
                        'blocked_paths': permission.blocked_paths or [],
                        'read_only_paths': permission.read_only_paths or [],
                        'frontend_url': settings.FRONTEND_URL,
                    }

                    html_content = render_to_string('emails/group_route_permissions_updated.html', context)

                    from django.core.mail import EmailMessage
                    from datetime import datetime

                    # Preparar nombre del archivo Excel
                    timestamp = datetime.now().strftime('%Y%m%d')
                    filename = f'{timestamp}_Permisos_Modificados_Grupo_{group_name}.xlsx'

                    email = EmailMessage(
                        subject=f'🔄 Permisos Modificados en Ruta - Grupo {group_name} - Sistema IGAC',
                        body=html_content,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[permission.user.email],
                    )
                    email.content_subtype = 'html'

                    # Adjuntar Excel del grupo completo
                    if excel_file:
                        excel_file.seek(0)  # Resetear puntero del archivo
                        email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                    email.send(fail_silently=False)

                    emails_sent.append(permission.user.email)
                    print(f"[OK] Email con Excel enviado a {permission.user.email}")

                except Exception as e:
                    print(f"[ERROR] Error enviando email a {permission.user.email}: {str(e)}")

        # ENVIAR EMAIL AL LÍDER
        if notify_leader_flag and permissions.exists():
            sample_perm = permissions.first()
            if sample_perm.authorized_by_email:
                try:
                    route_display = f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{route_path_normalized}"

                    # Contar usuarios únicos en el grupo
                    unique_users = UserPermission.objects.filter(
                        group_name=group_name,
                        is_active=True
                    ).values('user').distinct().count()

                    context_leader = {
                        'leader_name': sample_perm.authorized_by_name or 'Líder',
                        'user_name': sample_perm.authorized_by_name or 'Líder',
                        'group_name': group_name,
                        'base_path': route_path_normalized,
                        'route_display': route_display,
                        'total_users': unique_users,
                        'can_read': sample_perm.can_read,
                        'can_write': sample_perm.can_write,
                        'can_delete': sample_perm.can_delete,
                        'can_create_directories': sample_perm.can_create_directories,
                        'blocked_paths': sample_perm.blocked_paths or [],
                        'read_only_paths': sample_perm.read_only_paths or [],
                        'frontend_url': settings.FRONTEND_URL,
                    }

                    html_content_leader = render_to_string('emails/permission_modified_leader.html', context_leader)

                    from django.core.mail import EmailMessage
                    from datetime import datetime

                    # Preparar nombre del archivo Excel
                    timestamp = datetime.now().strftime('%Y%m%d')
                    filename_leader = f'{timestamp}_Permisos_Modificados_Grupo_{group_name}_LIDER.xlsx'

                    email_leader = EmailMessage(
                        subject=f'📧 Copia: Permisos Modificados - Grupo {group_name} - Sistema IGAC',
                        body=html_content_leader,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[sample_perm.authorized_by_email],
                    )
                    email_leader.content_subtype = 'html'

                    # Adjuntar Excel del grupo completo
                    if excel_file:
                        excel_file.seek(0)  # Resetear puntero del archivo
                        email_leader.attach(filename_leader, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                    email_leader.send(fail_silently=False)

                    emails_sent.append(sample_perm.authorized_by_email)
                    print(f"[OK] Email con Excel enviado al líder {sample_perm.authorized_by_email}")

                except Exception as e:
                    print(f"[ERROR] Error enviando email al líder: {str(e)}")

        return Response({
            'success': True,
            'group_name': group_name,
            'route_path': route_path,
            'updated_count': updated_count,
            'emails_sent': emails_sent,
            'message': f'Se actualizaron los permisos de la ruta para {updated_count} usuario(s)'
        })

    @action(detail=False, methods=['delete'], url_path=r'groups/(?P<group_name>[^/.]+)/routes/(?P<route_path>(?!.*\/permissions$).+)')
    def delete_route_from_group(self, request, group_name=None, route_path=None):
        """
        DELETE /api/admin/groups/{group_name}/routes/{route_path}/
        Elimina una ruta específica del grupo (elimina los permisos de TODOS los usuarios para esa ruta)

        Returns: {
            "success": true,
            "group_name": "...",
            "route_path": "...",
            "deleted_count": number,
            "users_affected": number,
            "message": "..."
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        # Decodificar la ruta (viene en URL encoding)
        from urllib.parse import unquote
        route_path = unquote(route_path)

        # Obtener permisos de esta ruta en el grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            base_path=route_path,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontraron permisos para la ruta "{route_path}" en el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        users_affected = permissions.values('user').distinct().count()
        deleted_count = 0

        try:
            with transaction.atomic():
                # Eliminar TODOS los permisos de esta ruta en el grupo
                deleted_count = permissions.delete()[0]

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='delete_file',
                    target_path=f'bulk_permissions/{group_name}/{route_path}',
                    target_name=f'Eliminar ruta de grupo: {group_name}',
                    details={
                        'action_type': 'delete_route_from_group',
                        'group_name': group_name,
                        'route_path': route_path,
                        'deleted_count': deleted_count,
                        'users_affected': users_affected
                    },
                    success=True
                )

                print(f"[OK] Ruta '{route_path}' eliminada del grupo '{group_name}' ({deleted_count} permisos eliminados, {users_affected} usuarios afectados)")

        except Exception as e:
            return Response(
                {'error': f'Error al eliminar ruta: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'route_path': route_path,
            'deleted_count': deleted_count,
            'users_affected': users_affected,
            'message': f'Ruta eliminada del grupo. Se eliminaron {deleted_count} permiso(s) de {users_affected} usuario(s)'
        })

    @action(detail=False, methods=['patch'], url_path='groups/(?P<group_name>[^/.]+)/expiration')
    def update_group_expiration(self, request, group_name=None):
        """
        PATCH /api/admin/groups/{group_name}/expiration/
        Actualiza la fecha de vencimiento de TODOS los permisos del grupo

        Body: {
            "expires_at": "2025-12-31T23:59:59Z"
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        expires_at = request.data.get('expires_at')

        if not expires_at:
            return Response(
                {'error': 'Debe proporcionar una fecha de vencimiento'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener todos los permisos del grupo
        permissions = UserPermission.objects.filter(
            group_name=group_name,
            is_active=True
        )

        if not permissions.exists():
            return Response(
                {'error': f'No se encontró el grupo "{group_name}"'},
                status=status.HTTP_404_NOT_FOUND
            )

        updated_count = 0

        try:
            with transaction.atomic():
                # Actualizar fecha de vencimiento de todos los permisos
                for permission in permissions:
                    permission.expires_at = expires_at
                    permission.save()
                    updated_count += 1

                # Registrar en auditoría
                from audit.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    username=request.user.username,
                    user_role=request.user.role,
                    action='create_folder',
                    target_path=f'bulk_permissions/{group_name}',
                    target_name=f'Actualizar vencimiento del grupo: {group_name}',
                    details={
                        'action_type': 'update_group_expiration',
                        'group_name': group_name,
                        'new_expires_at': str(expires_at),
                        'updated_count': updated_count
                    },
                    success=True
                )

        except Exception as e:
            return Response(
                {'error': f'Error al actualizar fecha de vencimiento: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            'success': True,
            'group_name': group_name,
            'updated_count': updated_count,
            'new_expires_at': expires_at,
            'message': f'Se actualizó la fecha de vencimiento de {updated_count} permiso(s)'
        })

    @action(detail=False, methods=['post'], url_path='generate-permissions-report')
    def generate_permissions_report(self, request):
        """
        POST /api/admin/users/generate-permissions-report/
        Genera un reporte CSV con los permisos de los usuarios seleccionados

        Body: {
            "user_ids": [1, 2, 3, ...]
        }

        Retorna un archivo CSV con todas las rutas y permisos de los usuarios,
        incluyendo permisos activos e inactivos para auditoría completa.
        """
        try:
            error = self._check_superadmin(request.user)
            if error:
                return error

            user_ids = request.data.get('user_ids', [])

            if not user_ids:
                return Response(
                    {'error': 'Debe proporcionar al menos un usuario'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Obtener usuarios
            users = User.objects.filter(id__in=user_ids).order_by('username')

            if not users.exists():
                return Response(
                    {'error': 'No se encontraron usuarios con los IDs proporcionados'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Obtener solo los permisos ACTIVOS de estos usuarios
            permissions = UserPermission.objects.filter(
                user_id__in=user_ids,
                is_active=True
            ).select_related('user').order_by('user__username', 'base_path')

            if not permissions.exists():
                return Response(
                    {'error': 'Los usuarios seleccionados no tienen permisos asignados'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Generar CSV
            import csv
            from django.http import HttpResponse

            # Crear respuesta HTTP con UTF-8 BOM para compatibilidad con Excel
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_permisos_{timestamp}.csv'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # UTF-8 BOM para que Excel reconozca caracteres hispanos
            response.write('\ufeff')

            writer = csv.writer(response)

            # Encabezados
            headers = [
                'Usuario',
                'Nombre Completo',
                'Correo',
                'Ruta Asignada',
                'Fecha Asignación',
                'Fecha Caducidad',
                'Estado',
                'Lectura',
                'Escritura',
                'Eliminación',
                'Crear Carpetas',
                'Exento Diccionario',
                'Nivel Edición',
                'Modo Herencia',
                'Profundidad Máxima',
                'Rutas Bloqueadas',
                'Rutas Solo Lectura',
                'Nombre Grupo'
            ]
            writer.writerow(headers)

            # Datos
            for perm in permissions:
                # Formatear fechas
                fecha_asignacion = perm.granted_at.strftime('%Y-%m-%d %H:%M:%S') if perm.granted_at else 'N/A'
                fecha_caducidad = perm.expires_at.strftime('%Y-%m-%d') if perm.expires_at else 'Sin vencimiento'

                # Estado
                estado = 'Activo' if perm.is_active else 'Inactivo'

                # Permisos booleanos
                lectura = 'Sí' if perm.can_read else 'No'
                escritura = 'Sí' if perm.can_write else 'No'
                eliminacion = 'Sí' if perm.can_delete else 'No'
                crear_carpetas = 'Sí' if perm.can_create_directories else 'No'
                exento_diccionario = 'Sí' if perm.exempt_from_dictionary else 'No'

                # Nivel de edición
                nivel_map = {
                    'upload_only': 'Solo Subir',
                    'upload_own': 'Subir + Editar Propios',
                    'upload_all': 'Subir + Editar Todos',
                    '': 'Sin restricción',
                    None: 'Sin restricción'
                }
                nivel_edicion = nivel_map.get(perm.edit_permission_level, perm.edit_permission_level or 'Sin restricción')

                # Modo de herencia
                modo_map = {
                    'total': 'Herencia Total',
                    'blocked': 'Herencia con Bloqueos',
                    'limited_depth': 'Profundidad Limitada',
                    'partial_write': 'Herencia Parcial de Escritura',
                    'none': 'Sin Herencia',
                    '': 'Herencia Total',
                    None: 'Herencia Total'
                }
                modo_herencia = modo_map.get(perm.inheritance_mode, perm.inheritance_mode or 'Herencia Total')

                # Profundidad máxima
                profundidad_maxima = str(perm.max_depth) if perm.max_depth is not None else 'Ilimitada'

                # Rutas bloqueadas y solo lectura (convertir lista a string)
                rutas_bloqueadas = ', '.join(perm.blocked_paths) if perm.blocked_paths else 'Ninguna'
                rutas_solo_lectura = ', '.join(perm.read_only_paths) if perm.read_only_paths else 'Ninguna'

                # Nombre del grupo
                nombre_grupo = perm.group_name or 'Sin grupo'

                # Construir ruta completa para CSV (normalizar slashes a backslash)
                if perm.base_path:
                    # Reemplazar forward slashes por backslashes para rutas de Windows
                    base_path_normalizada = perm.base_path.replace('/', '\\')
                    base_path_completa = f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_normalizada}"
                else:
                    base_path_completa = "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy"

                # Escribir fila
                row = [
                    perm.user.username,
                    f"{perm.user.first_name} {perm.user.last_name}".strip() or 'N/A',
                    perm.user.email,
                    base_path_completa,
                    fecha_asignacion,
                    fecha_caducidad,
                    estado,
                    lectura,
                    escritura,
                    eliminacion,
                    crear_carpetas,
                    exento_diccionario,
                    nivel_edicion,
                    modo_herencia,
                    profundidad_maxima,
                    rutas_bloqueadas,
                    rutas_solo_lectura,
                    nombre_grupo
                ]
                writer.writerow(row)

            # Registrar en auditoría
            from audit.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                username=request.user.username,
                user_role=request.user.role,
                action='download',
                target_path='admin/reports',
                target_name=filename,
                details={
                    'action_type': 'generate_permissions_report',
                    'user_ids': user_ids,
                    'user_count': len(user_ids),
                    'permission_count': permissions.count()
                },
                success=True
            )

            return response
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"ERROR EN generate_permissions_report: {str(e)}")
            print(error_trace)
            return Response(
                {'error': f'Error generando reporte: {str(e)}', 'trace': error_trace},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generate_excel_report(self, permissions, authorized_by_name):
        """
        Genera un reporte Excel profesional con permisos
        FORMATO: UNA FILA POR CADA RUTA (cada permiso individual)

        Args:
            permissions: QuerySet de UserPermission a incluir en el reporte
            authorized_by_name: Nombre completo de quien autorizó

        Returns:
            BytesIO: Archivo Excel en memoria listo para descarga o adjuntar en email
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        from io import BytesIO
        from datetime import datetime

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Permisos Repositorio"

        # ==================== ESTILOS PROFESIONALES ====================

        # Título principal
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_alignment = Alignment(horizontal='center', vertical='center')

        # Encabezados de tabla
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Bordes
        thick_border = Border(
            left=Side(style='medium', color='1F4E78'),
            right=Side(style='medium', color='1F4E78'),
            top=Side(style='medium', color='1F4E78'),
            bottom=Side(style='medium', color='1F4E78')
        )
        thin_border = Border(
            left=Side(style='thin', color='8EA9DB'),
            right=Side(style='thin', color='8EA9DB'),
            top=Side(style='thin', color='8EA9DB'),
            bottom=Side(style='thin', color='8EA9DB')
        )

        # Celdas de datos
        cell_font = Font(name='Arial', size=10, color="000000")
        cell_font_bold = Font(name='Arial', size=10, color="000000", bold=True)
        cell_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Colores para permisos
        green_font = Font(name='Arial', size=10, color="008000", bold=True)  # Verde para permisos normales
        red_font = Font(name='Arial', size=10, color="FF0000", bold=True)    # Rojo para eliminación

        # Colores alternados para filas (zebra striping)
        row_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_light_blue = PatternFill(start_color="F2F5FA", end_color="F2F5FA", fill_type="solid")

        # ==================== ENCABEZADO DEL REPORTE ====================

        current_row = 1

        # Título principal (fusionar columnas A-G)
        ws.merge_cells(f'A{current_row}:G{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "INSTITUTO GEOGRÁFICO AGUSTÍN CODAZZI - IGAC"
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        title_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Subtítulo del reporte
        ws.merge_cells(f'A{current_row}:G{current_row}')
        subtitle_cell = ws[f'A{current_row}']
        subtitle_cell.value = "REPORTE DE PERMISOS - REPOSITORIO SUBDIRECCION DE PROYECTOS"
        subtitle_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subtitle_cell.font = Font(name='Arial', size=13, bold=True, color="1F4E78")
        subtitle_cell.alignment = title_alignment
        subtitle_cell.border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Fila vacía
        current_row += 1

        # Información del reporte
        info_font = Font(name='Arial', size=10, bold=True, color="1F4E78")
        info_value_font = Font(name='Arial', size=10, color="000000")

        # Fecha de generación
        ws[f'A{current_row}'] = "Fecha de generación:"
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = datetime.now().strftime("%d de %B de %Y a las %H:%M")
        ws[f'B{current_row}'].font = info_value_font
        ws.merge_cells(f'B{current_row}:D{current_row}')
        current_row += 1

        # Solicitado por
        if authorized_by_name:
            ws[f'A{current_row}'] = "Solicitado por:"
            ws[f'A{current_row}'].font = info_font
            ws[f'B{current_row}'] = authorized_by_name
            ws[f'B{current_row}'].font = info_value_font
            ws.merge_cells(f'B{current_row}:D{current_row}')
            current_row += 1

        # Total de usuarios
        num_usuarios = len(set(perm.user.id for perm in permissions))
        ws[f'A{current_row}'] = "Total de usuarios:"
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = num_usuarios
        ws[f'B{current_row}'].font = info_value_font
        current_row += 1

        # Fila vacía antes de la tabla
        current_row += 1

        # ==================== TABLA DE DATOS ====================

        # Encabezados de la tabla
        headers = ['Nº', 'USUARIO', 'NOMBRE COMPLETO', 'GRUPO/PROYECTO', 'RUTA DE ACCESO AUTORIZADA', 'PERMISOS ASIGNADOS', 'VIGENCIA']
        header_row = current_row

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thick_border

        ws.row_dimensions[header_row].height = 35
        current_row += 1

        # ==================== GENERAR FILAS DE DATOS (UNA POR RUTA) ====================

        row_counter = 1
        # Manejar tanto QuerySets como listas de permisos
        if hasattr(permissions, 'select_related'):
            # Es un QuerySet
            perms_to_iterate = permissions.select_related('user').order_by('user__email', 'base_path')
        else:
            # Es una lista
            perms_to_iterate = sorted(permissions, key=lambda p: (p.user.email, p.base_path))

        for perm in perms_to_iterate:
            row_num = current_row

            # Construir ruta completa y normalizada
            if perm.base_path:
                base_path_normalizada = perm.base_path.replace('/', '\\')
                ruta_completa = f"\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path_normalizada}"
            else:
                ruta_completa = "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy"

            # Formato de vigencia
            if perm.expires_at:
                vigencia_texto = perm.expires_at.strftime('%d/%m/%Y')
                dias_restantes = (perm.expires_at.date() - datetime.now().date()).days
                if dias_restantes <= 7:
                    vigencia_texto += f"\n⚠ {dias_restantes} días"
            else:
                vigencia_texto = "Sin vencimiento"

            # Datos básicos de la fila
            row_data = [
                row_counter,  # Nº
                perm.user.email,  # USUARIO
                perm.user.get_full_name() or perm.user.username,  # NOMBRE COMPLETO
                perm.group_name or 'Sin grupo',  # GRUPO/PROYECTO
                ruta_completa,  # RUTA DE ACCESO
                '',  # PERMISOS (se llenará después con formato especial)
                vigencia_texto  # VIGENCIA
            ]

            # Escribir las columnas básicas (A-E, G)
            for col_num in [1, 2, 3, 4, 5, 7]:
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = row_data[col_num - 1]
                cell.border = thin_border

                # Aplicar zebra striping
                if row_counter % 2 == 0:
                    cell.fill = row_light_blue
                else:
                    cell.fill = row_white

                # Aplicar alineación según columna
                if col_num == 1:  # Nº: centrado y negrilla
                    cell.font = cell_font_bold
                    cell.alignment = cell_alignment_center
                elif col_num in [2, 3, 4, 7]:  # Usuario, Nombre, Grupo, Vigencia: centrado
                    cell.font = cell_font
                    cell.alignment = cell_alignment_center
                else:  # Ruta: izquierda
                    cell.font = cell_font
                    cell.alignment = cell_alignment_left

            # COLUMNA F (PERMISOS): Texto con colores según tipo
            permisos_cell = ws.cell(row=row_num, column=6)
            permisos_cell.border = thin_border
            permisos_cell.alignment = cell_alignment_left
            if row_counter % 2 == 0:
                permisos_cell.fill = row_light_blue
            else:
                permisos_cell.fill = row_white

            # Construir texto de permisos con colores (AZUL para todo, ROJO solo para Eliminación)
            text_blocks = []
            blue_font = InlineFont(color='0000FF', b=True, rFont='Arial', sz=10)  # Azul
            red_font_inline = InlineFont(color='FF0000', b=True, rFont='Arial', sz=10)  # Rojo

            if perm.can_read:
                text_blocks.append(TextBlock(blue_font, "✓ Lectura\n"))
            if perm.can_write:
                text_blocks.append(TextBlock(blue_font, "✓ Escritura\n"))
            if perm.can_delete:
                text_blocks.append(TextBlock(red_font_inline, "✓ Eliminación\n"))  # ROJO
            if perm.can_create_directories:
                text_blocks.append(TextBlock(blue_font, "✓ Crear Carpetas\n"))
            if perm.exempt_from_dictionary:
                text_blocks.append(TextBlock(blue_font, "✓ Exento Diccionario\n"))

            if perm.edit_permission_level:
                nivel_map = {
                    'upload_only': 'Solo Subir',
                    'upload_own': 'Subir + Editar Propios',
                    'upload_all': 'Subir + Editar Todos'
                }
                text_blocks.append(TextBlock(blue_font, f"• Nivel: {nivel_map.get(perm.edit_permission_level, perm.edit_permission_level)}"))

            if text_blocks:
                permisos_cell.value = CellRichText(*text_blocks)
            else:
                permisos_cell.value = 'Sin permisos'
                permisos_cell.font = Font(name='Arial', size=10, color="666666")

            # Altura de fila (suficiente para mostrar todos los permisos)
            ws.row_dimensions[row_num].height = 90

            current_row += 1
            row_counter += 1

        # ==================== AJUSTAR ANCHOS DE COLUMNA ====================

        column_widths = {
            'A': 28,  # Nº (ancho suficiente para no tapar "Fecha de generación:", "Solicitado por:", etc.)
            'B': 30,  # USUARIO (email)
            'C': 28,  # NOMBRE COMPLETO
            'D': 25,  # GRUPO/PROYECTO
            'E': 70,  # RUTA DE ACCESO (más angosto porque solo hay 1 ruta por fila)
            'F': 30,  # PERMISOS
            'G': 15   # VIGENCIA (centrado)
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ==================== PIE DEL REPORTE ====================

        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        footer_cell = ws[f'A{current_row}']
        footer_cell.value = "Este documento es generado automáticamente por el Sistema de Gestión de Repositorio NetApp - IGAC"
        footer_cell.font = Font(name='Arial', size=8, italic=True, color="666666")
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20

        # Congelar panel (mantener encabezados visibles al hacer scroll)
        ws.freeze_panes = f'A{header_row + 1}'

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def _generate_group_excel_report_by_route(self, permissions, authorized_by_name, group_name):
        """
        Genera un reporte Excel AGRUPADO POR RUTA para reportes de grupo.
        Muestra cada ruta UNA SOLA VEZ con todos los usuarios que tienen acceso a ella.

        Args:
            permissions: QuerySet de UserPermission del grupo
            authorized_by_name: Nombre completo de quien autorizó
            group_name: Nombre del grupo

        Returns:
            BytesIO: Archivo Excel en memoria
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        from io import BytesIO
        from datetime import datetime
        from collections import defaultdict

        # Agrupar permisos por base_path (ruta)
        routes_dict = defaultdict(list)
        for perm in permissions:
            routes_dict[perm.base_path].append(perm)

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Permisos Repositorio"

        # ==================== ESTILOS PROFESIONALES ====================

        # Título principal
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        title_alignment = Alignment(horizontal='center', vertical='center')

        # Subtítulo
        subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subtitle_font = Font(name='Arial', size=13, bold=True, color="FFFFFF")

        # Encabezados de tabla
        header_fill = PatternFill(start_color="2E5C9A", end_color="2E5C9A", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Alineaciones
        cell_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Colores de filas alternadas
        row_light_blue = PatternFill(start_color="D9E9F7", end_color="D9E9F7", fill_type="solid")
        row_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ==================== TÍTULO Y ENCABEZADO ====================

        # Título principal
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = 'INSTITUTO GEOGRÁFICO AGUSTÍN CODAZZI - IGAC'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 30

        # Subtítulo
        ws.merge_cells('A2:C2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = 'REPORTE DE PERMISOS - REPOSITORIO SUBDIRECCION DE PROYECTOS'
        subtitle_cell.fill = subtitle_fill
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = title_alignment
        ws.row_dimensions[2].height = 25

        # Información del reporte
        current_row = 4
        info_font = Font(name='Arial', size=10, bold=True)

        ws[f'A{current_row}'] = 'Fecha de generación:'
        ws[f'B{current_row}'] = datetime.now().strftime('%d de %B de %Y a las %H:%M')
        ws[f'A{current_row}'].font = info_font

        current_row += 1
        ws[f'A{current_row}'] = 'Solicitado por:'
        ws[f'B{current_row}'] = authorized_by_name
        ws[f'A{current_row}'].font = info_font

        current_row += 1
        ws[f'A{current_row}'] = 'Grupo/Proyecto:'
        ws[f'B{current_row}'] = group_name
        ws[f'A{current_row}'].font = info_font

        current_row += 1
        ws[f'A{current_row}'] = 'Total de rutas:'
        ws[f'B{current_row}'] = len(routes_dict)
        ws[f'A{current_row}'].font = info_font

        # USUARIOS CON ACCESO (listar nombres en una fila)
        current_row += 1
        ws[f'A{current_row}'] = 'Usuarios con acceso:'
        ws[f'A{current_row}'].font = info_font
        all_users_names = []
        all_users_emails = []
        for perms_list in routes_dict.values():
            for perm in perms_list:
                name = perm.user.get_full_name()
                email = perm.user.email
                if name not in all_users_names:
                    all_users_names.append(name)
                if email not in all_users_emails:
                    all_users_emails.append(email)

        # Nombres en esta fila
        ws[f'B{current_row}'] = ', '.join(all_users_names)

        # CORREOS ELECTRÓNICOS (en la siguiente fila)
        current_row += 1
        ws[f'A{current_row}'] = 'Correos electrónicos:'
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = ', '.join(all_users_emails)

        # PERMISOS ASIGNADOS (mostrar una sola vez con colores)
        current_row += 1
        ws[f'A{current_row}'] = 'Permisos asignados:'
        ws[f'A{current_row}'].font = info_font
        first_perm = list(routes_dict.values())[0][0]

        # Crear TextBlocks con colores (azul para normal, rojo para eliminación)
        blue_font = InlineFont(color='0000FF', b=True, rFont='Arial', sz=10)
        red_font = InlineFont(color='FF0000', b=True, rFont='Arial', sz=10)

        text_blocks = []
        if first_perm.can_read:
            text_blocks.append(TextBlock(blue_font, "Lectura, "))
        if first_perm.can_write:
            text_blocks.append(TextBlock(blue_font, "Escritura, "))
        if first_perm.can_delete:
            text_blocks.append(TextBlock(red_font, "Eliminación, "))  # ROJO
        if first_perm.can_create_directories:
            text_blocks.append(TextBlock(blue_font, "Crear Carpetas, "))
        if first_perm.exempt_from_dictionary:
            text_blocks.append(TextBlock(blue_font, "Exento Diccionario, "))
        if first_perm.edit_permission_level:
            nivel_map = {
                'upload_only': 'Solo Subir',
                'upload_own': 'Subir + Editar Propios',
                'upload_all': 'Subir + Editar Todos'
            }
            nivel_text = f"Nivel: {nivel_map.get(first_perm.edit_permission_level, first_perm.edit_permission_level)}, "
            text_blocks.append(TextBlock(blue_font, nivel_text))

        # Remover última coma
        if text_blocks and text_blocks[-1].text.endswith(', '):
            last_block = text_blocks[-1]
            text_blocks[-1] = TextBlock(last_block.font, last_block.text[:-2])

        if text_blocks:
            ws[f'B{current_row}'] = CellRichText(*text_blocks)
        else:
            ws[f'B{current_row}'] = 'Sin permisos'

        # ==================== ENCABEZADOS DE TABLA ====================

        current_row += 2
        header_row = current_row
        headers = ['Nº', 'RUTA DE ACCESO AUTORIZADA', 'VIGENCIA']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[header_row].height = 40

        # ==================== DATOS AGRUPADOS POR RUTA ====================

        current_row += 1
        row_counter = 1

        # Ordenar rutas alfabéticamente
        sorted_routes = sorted(routes_dict.items())

        for base_path, perms_for_route in sorted_routes:
            row_num = current_row

            # Tomar el primer permiso como referencia (todos deberían tener mismos permisos/vigencia)
            first_perm = perms_for_route[0]

            # Columna 1: Nº
            num_cell = ws.cell(row=row_num, column=1)
            num_cell.value = row_counter
            num_cell.font = Font(name='Arial', size=10, bold=True)
            num_cell.alignment = cell_alignment_center
            num_cell.border = thin_border
            if row_counter % 2 == 0:
                num_cell.fill = row_light_blue
            else:
                num_cell.fill = row_white

            # Columna 2: RUTA DE ACCESO
            ruta_cell = ws.cell(row=row_num, column=2)
            ruta_cell.value = f'\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\{base_path}'
            ruta_cell.font = Font(name='Arial', size=10)
            ruta_cell.alignment = cell_alignment_left
            ruta_cell.border = thin_border
            if row_counter % 2 == 0:
                ruta_cell.fill = row_light_blue
            else:
                ruta_cell.fill = row_white

            # Columna 3: VIGENCIA
            vigencia_cell = ws.cell(row=row_num, column=3)
            if first_perm.expires_at:
                vigencia_cell.value = first_perm.expires_at.strftime('%d/%m/%Y')
            else:
                vigencia_cell.value = 'Indefinido'
            vigencia_cell.font = Font(name='Arial', size=10)
            vigencia_cell.alignment = cell_alignment_center
            vigencia_cell.border = thin_border
            if row_counter % 2 == 0:
                vigencia_cell.fill = row_light_blue
            else:
                vigencia_cell.fill = row_white

            # Altura de fila
            ws.row_dimensions[row_num].height = 25

            current_row += 1
            row_counter += 1

        # ==================== AJUSTAR ANCHOS DE COLUMNA ====================

        column_widths = {
            'A': 22,  # Nº (más ancho como solicitó el usuario)
            'B': 80,  # RUTA DE ACCESO (más ancho)
            'C': 15   # VIGENCIA
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ==================== PIE DEL REPORTE ====================

        current_row += 1
        footer_cell = ws.cell(row=current_row, column=1)
        footer_cell.value = f'Documento generado automáticamente el {datetime.now().strftime("%d/%m/%Y a las %H:%M")} - Sistema de Gestión de Archivos NetApp IGAC'
        ws.merge_cells(f'A{current_row}:C{current_row}')
        footer_cell.font = Font(name='Arial', size=8, italic=True, color="666666")
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20

        # Congelar panel
        ws.freeze_panes = f'A{header_row + 1}'

        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def _generate_secure_password(self, length=12):
        """Genera una contraseña segura de 12 caracteres"""
        uppercase = string.ascii_uppercase
        lowercase = string.ascii_lowercase
        digits = string.digits
        symbols = '!@#$%^&*()_+-=[]{}|;:,.<>?/~'
        all_chars = uppercase + lowercase + digits + symbols

        # Asegurar al menos un carácter de cada tipo
        password = [
            secrets.choice(uppercase),
            secrets.choice(lowercase),
            secrets.choice(digits),
            secrets.choice(symbols),
        ]

        # Completar hasta la longitud deseada
        password += [secrets.choice(all_chars) for _ in range(length - 4)]

        # Mezclar
        secrets.SystemRandom().shuffle(password)

        return ''.join(password)

    def _send_welcome_email(self, user, password, is_reset=False):
        """Envía email de bienvenida con credenciales"""
        subject = 'Nueva cuenta - NetApp Bridge IGAC' if not is_reset else 'Credenciales actualizadas - NetApp Bridge IGAC'

        # Renderizar template HTML
        html_message = render_to_string('emails/welcome.html', {
            'user': user,
            'password': password,
            'is_reset': is_reset,
            'login_url': f'{settings.FRONTEND_URL}/login' if hasattr(settings, 'FRONTEND_URL') else 'http://localhost:4545/login'
        })

        # Texto plano (fallback)
        plain_message = strip_tags(html_message)

        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )

    @action(detail=False, methods=['get'], url_path='path-access/(?P<path_param>.*)')
    def get_path_access(self, request, path_param=None):
        """
        GET /api/admin/users/path-access/{path}/
        Obtiene todos los usuarios que tienen acceso a una ruta específica

        Returns: {
            "path": "05_grup_trab/11_gest_info",
            "users_with_access": [
                {
                    "user": {...},
                    "permission": {...},
                    "access_type": "direct" | "inherited",
                    "effective_permissions": {
                        "can_read": true,
                        "can_write": false,
                        ...
                    }
                }
            ]
        }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        try:
            # Decodificar la ruta del parámetro URL
            from urllib.parse import unquote
            path = unquote(path_param) if path_param else ''

            # Obtener todos los permisos activos
            all_permissions = UserPermission.objects.filter(
                is_active=True
            ).select_related('user').order_by('user__username')

            users_with_access = []

            for perm in all_permissions:
                # Usar PermissionService para verificar si este permiso da acceso a la ruta
                from services.permission_service import PermissionService

                # Verificar si la ruta está dentro del scope del permiso
                normalized_path = PermissionService.normalize_path(path)
                normalized_base = PermissionService.normalize_path(perm.base_path)

                # Determinar si hay acceso
                has_access = False
                access_type = None

                # Caso 1: Permiso directo (base_path == path)
                if normalized_base == normalized_path:
                    has_access = True
                    access_type = 'direct'

                # Caso 2: Permiso heredado (path está dentro de base_path)
                elif normalized_path.startswith(normalized_base + '/') or normalized_base == '':
                    # Verificar que no esté bloqueado
                    if not PermissionService.is_path_blocked(perm, normalized_path):
                        # Verificar profundidad máxima si aplica
                        if perm.inheritance_mode == 'limited_depth' and perm.max_depth is not None:
                            depth = PermissionService.get_path_depth(normalized_base, normalized_path)
                            if depth <= perm.max_depth:
                                has_access = True
                                access_type = 'inherited'
                        elif perm.inheritance_mode in ['total', 'blocked', 'partial_write']:
                            has_access = True
                            access_type = 'inherited'

                if has_access:
                    # Determinar permisos efectivos
                    effective_perms = {
                        'can_read': perm.can_read,
                        'can_write': perm.can_write,
                        'can_delete': perm.can_delete,
                        'can_create_directories': perm.can_create_directories,
                    }

                    # Si es partial_write, verificar si la ruta está en read_only_paths
                    if perm.inheritance_mode == 'partial_write':
                        read_only_paths = perm.read_only_paths or []
                        for ro_path in read_only_paths:
                            ro_normalized = PermissionService.normalize_path(ro_path)
                            if normalized_path == ro_normalized or normalized_path.startswith(ro_normalized + '/'):
                                effective_perms['can_write'] = False
                                effective_perms['can_delete'] = False
                                effective_perms['can_create_directories'] = False
                                break

                    users_with_access.append({
                        'user': UserSerializer(perm.user).data,
                        'permission': UserPermissionSerializer(perm).data,
                        'access_type': access_type,
                        'effective_permissions': effective_perms
                    })

            return Response({
                'path': path,
                'users_with_access': users_with_access,
                'total_users': len(users_with_access)
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error al obtener accesos: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _check_parent_permissions(self, user, target_path):
        """
        Verifica si el usuario tiene permisos en rutas SUPERIORES (padres) de la ruta objetivo.

        Args:
            user: Usuario a verificar
            target_path: Ruta objetivo (ej: "05_grup_trab/11_gest_info/2025/datos")

        Returns:
            dict con información del permiso padre si existe, None si no existe

        Ejemplo:
            Si el usuario tiene permisos en "05_grup_trab/11_gest_info" (padre)
            y se verifica "05_grup_trab/11_gest_info/2025/datos" (hijo),
            retorna información del permiso padre.
        """
        if not target_path or target_path == '':
            return None

        # Normalizar path (convertir a forward slashes)
        target_path_normalized = target_path.replace('\\', '/')

        # Obtener todos los permisos activos del usuario
        user_permissions = UserPermission.objects.filter(
            user=user,
            is_active=True
        )

        for perm in user_permissions:
            parent_path = perm.base_path or ''
            parent_path_normalized = parent_path.replace('\\', '/')

            # Verificar si el permiso es de una ruta superior (padre)
            # Caso 1: Permiso en raíz (base_path vacío o None) - cubre TODO
            if not parent_path_normalized or parent_path_normalized == '':
                # Usuario tiene permiso en la raíz - cubre todas las rutas
                restriction_summary = []
                if perm.blocked_paths:
                    restriction_summary.append(f"{len(perm.blocked_paths)} ruta(s) bloqueada(s)")
                if perm.read_only_paths:
                    restriction_summary.append(f"{len(perm.read_only_paths)} ruta(s) solo lectura")
                if perm.max_depth is not None:
                    restriction_summary.append(f"profundidad máxima: {perm.max_depth}")

                return {
                    'id': perm.id,
                    'base_path': parent_path,
                    'base_path_display': 'Sub_Proy (raíz)',
                    'blocked_paths': perm.blocked_paths,
                    'read_only_paths': perm.read_only_paths,
                    'max_depth': perm.max_depth,
                    'restriction_summary': ', '.join(restriction_summary) if restriction_summary else None
                }

            # Caso 2: target_path empieza con parent_path (es subdirectorio)
            # Ejemplo: parent="05_grup_trab/11_gest_info", target="05_grup_trab/11_gest_info/2025"
            if target_path_normalized.startswith(parent_path_normalized + '/'):
                # Verificar si la ruta objetivo está bloqueada
                is_blocked = False
                if perm.blocked_paths:
                    for blocked in perm.blocked_paths:
                        blocked_normalized = blocked.replace('\\', '/')
                        # Construir ruta completa del bloqueo
                        full_blocked_path = f"{parent_path_normalized}/{blocked_normalized}" if parent_path_normalized else blocked_normalized

                        if target_path_normalized == full_blocked_path or target_path_normalized.startswith(full_blocked_path + '/'):
                            is_blocked = True
                            break

                # Si está bloqueada, no cuenta como permiso padre válido
                if is_blocked:
                    continue

                restriction_summary = []
                if perm.blocked_paths:
                    restriction_summary.append(f"{len(perm.blocked_paths)} ruta(s) bloqueada(s)")
                if perm.read_only_paths:
                    restriction_summary.append(f"{len(perm.read_only_paths)} ruta(s) solo lectura")
                if perm.max_depth is not None:
                    restriction_summary.append(f"profundidad máxima: {perm.max_depth}")

                return {
                    'id': perm.id,
                    'base_path': parent_path,
                    'base_path_display': f'Sub_Proy/{parent_path}',
                    'blocked_paths': perm.blocked_paths,
                    'read_only_paths': perm.read_only_paths,
                    'max_depth': perm.max_depth,
                    'restriction_summary': ', '.join(restriction_summary) if restriction_summary else None
                }

        return None

    @action(detail=False, methods=['post'], url_path='validate-paths')
    def validate_paths(self, request):
        """
        POST /api/admin/users/validate-paths
        Valida la existencia de múltiples rutas en la NetApp y verifica si usuarios ya tienen permisos

        Request body:
            {
                "paths": ["05_grup_trab/11_gest_info", "03_gest_info_catas/2025"],
                "user_id": 123,        # Opcional: verificar un solo usuario (asignación individual)
                "user_ids": [1, 2, 3]  # Opcional: verificar múltiples usuarios (asignación masiva)
            }

        Response:
            {
                "results": [
                    {
                        "path": "05_grup_trab/11_gest_info",
                        "exists": true,
                        "full_path": "\\\\repositorio\\DirGesCat\\2510SP\\H_Informacion_Consulta\\Sub_Proy\\05_grup_trab\\11_gest_info",
                        "is_directory": true,
                        "has_permission": true,  # Si user_id fue proporcionado (individual)
                        "warning": "El usuario ya tiene permisos activos en esta ruta",
                        "users_with_permission": [  # Si user_ids fue proporcionado (masivo)
                            {"id": 1, "email": "user1@igac.gov.co", "name": "Usuario 1"},
                            {"id": 2, "email": "user2@igac.gov.co", "name": "Usuario 2"}
                        ]
                    },
                    {
                        "path": "ruta/inexistente",
                        "exists": false,
                        "error": "La ruta no existe"
                    }
                ]
            }
        """
        error = self._check_superadmin(request.user)
        if error:
            return error

        paths = request.data.get('paths', [])
        user_id = request.data.get('user_id')       # Para asignación individual
        user_ids = request.data.get('user_ids', [])  # Para asignación masiva

        if not paths or not isinstance(paths, list):
            return Response(
                {'error': 'Se requiere un array de rutas en el campo "paths"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Si se proporciona user_id (individual), obtener el usuario
        user = None
        if user_id:
            try:
                user = User.objects.get(pk=user_id)
            except User.DoesNotExist:
                return Response(
                    {'error': 'Usuario no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

        # Si se proporciona user_ids (masivo), obtener los usuarios
        users = []
        if user_ids and isinstance(user_ids, list):
            users = User.objects.filter(pk__in=user_ids)

        from services.smb_service import SMBService
        from services.path_converter import PathConverter

        smb_service = SMBService()
        results = []

        for path in paths:
            try:
                # Detectar y convertir rutas Windows UNC a rutas relativas
                original_path = path
                if PathConverter.is_windows_path(path):
                    # Convertir ruta Windows UNC a Linux completa
                    linux_full_path = PathConverter.windows_to_linux(path)
                    # Extraer solo la parte relativa (después del mount point)
                    if linux_full_path.startswith(PathConverter.LINUX_MOUNT_POINT):
                        path = linux_full_path[len(PathConverter.LINUX_MOUNT_POINT):].lstrip('/')
                    else:
                        path = linux_full_path.lstrip('/')

                # Construir ruta completa
                full_path = smb_service.build_full_path(path)

                # Verificar si existe
                exists = smb_service.file_exists(path)

                # Convertir full_path a formato Windows UNC para display
                full_path_display = PathConverter.linux_to_windows(full_path)

                result = {
                    'path': original_path,  # Mantener la ruta original para el frontend
                    'relative_path': path,  # Ruta relativa para uso interno
                    'exists': exists,
                    'full_path': full_path_display  # Mostrar en formato Windows UNC
                }

                if exists:
                    # Verificar si es directorio
                    import os
                    result['is_directory'] = os.path.isdir(full_path)
                else:
                    result['error'] = 'La ruta no existe'

                # CASO 1: Verificación individual (user_id proporcionado)
                if user:
                    # Verificar si tiene permiso EXACTO en esta ruta
                    existing_permission = UserPermission.objects.filter(
                        user=user,
                        base_path=path,
                        is_active=True
                    ).first()

                    # Verificar si tiene permisos en rutas SUPERIORES (padres)
                    parent_permission = self._check_parent_permissions(user, path)

                    if parent_permission:
                        # Usuario tiene permisos en ruta superior
                        has_restrictions = (
                            parent_permission.get('blocked_paths') or
                            parent_permission.get('read_only_paths') or
                            parent_permission.get('max_depth') is not None
                        )

                        if has_restrictions:
                            result['has_permission'] = True
                            result['has_parent_permission'] = True
                            result['parent_permission_info'] = parent_permission
                            result['warning'] = f"⚠️ El usuario ya tiene permisos en '{parent_permission['base_path_display']}' (ruta superior) con RESTRICCIONES. Los nuevos permisos sobrescribirán las restricciones anteriores."
                            result['alert_type'] = 'parent_with_restrictions'
                        else:
                            result['has_permission'] = False
                            result['has_parent_permission'] = True
                            result['parent_permission_info'] = parent_permission
                            result['info'] = f"✓ El usuario NO necesita permisos aquí. Ya tiene acceso TOTAL desde '{parent_permission['base_path_display']}' (ruta superior) sin restricciones."
                            result['alert_type'] = 'parent_without_restrictions'
                    elif existing_permission:
                        result['has_permission'] = True
                        result['warning'] = 'El usuario ya tiene permisos activos en esta ruta exacta'
                        result['alert_type'] = 'exact_match'
                    else:
                        result['has_permission'] = False

                # CASO 2: Verificación masiva (user_ids proporcionado)
                elif users:
                    # Encontrar qué usuarios ya tienen permisos en esta ruta EXACTA
                    existing_permissions = UserPermission.objects.filter(
                        user__in=users,
                        base_path=path,
                        is_active=True
                    ).select_related('user')

                    # Verificar permisos en rutas SUPERIORES para cada usuario
                    users_with_permission = []
                    users_with_parent_permission = []

                    for u in users:
                        # Verificar permiso exacto
                        exact_perm = existing_permissions.filter(user=u).first()

                        # Verificar permiso en ruta superior
                        parent_perm = self._check_parent_permissions(u, path)

                        if parent_perm:
                            has_restrictions = (
                                parent_perm.get('blocked_paths') or
                                parent_perm.get('read_only_paths') or
                                parent_perm.get('max_depth') is not None
                            )

                            if has_restrictions:
                                users_with_parent_permission.append({
                                    'id': u.id,
                                    'email': u.email,
                                    'name': u.get_full_name(),
                                    'parent_path': parent_perm['base_path_display'],
                                    'has_restrictions': True,
                                    'restriction_details': parent_perm.get('restriction_summary', '')
                                })
                            else:
                                users_with_parent_permission.append({
                                    'id': u.id,
                                    'email': u.email,
                                    'name': u.get_full_name(),
                                    'parent_path': parent_perm['base_path_display'],
                                    'has_restrictions': False
                                })
                        elif exact_perm:
                            users_with_permission.append({
                                'id': u.id,
                                'email': u.email,
                                'name': u.get_full_name()
                            })

                    result['users_with_permission'] = users_with_permission
                    result['users_with_parent_permission'] = users_with_parent_permission

                    if users_with_parent_permission:
                        users_no_restriction = [u for u in users_with_parent_permission if not u['has_restrictions']]
                        users_with_restriction = [u for u in users_with_parent_permission if u['has_restrictions']]

                        if users_no_restriction and users_with_restriction:
                            result['warning'] = f"⚠️ {len(users_no_restriction)} usuario(s) NO necesitan permisos (ya tienen acceso total desde ruta superior). {len(users_with_restriction)} usuario(s) tienen restricciones que serán sobrescritas."
                        elif users_no_restriction:
                            result['info'] = f"✓ {len(users_no_restriction)} usuario(s) NO necesitan permisos aquí (ya tienen acceso total desde ruta superior)"
                        elif users_with_restriction:
                            result['warning'] = f"⚠️ {len(users_with_restriction)} usuario(s) tienen permisos en ruta superior CON restricciones que serán sobrescritas"

                    if users_with_permission:
                        existing_msg = f"{len(users_with_permission)} usuario(s) ya tienen permisos en esta ruta exacta"
                        if result.get('warning'):
                            result['warning'] = f"{existing_msg}. {result['warning']}"
                        elif result.get('info'):
                            result['warning'] = existing_msg
                        else:
                            result['warning'] = existing_msg

                results.append(result)

            except Exception as e:
                results.append({
                    'path': path,
                    'exists': False,
                    'error': str(e)
                })

        return Response({'results': results})

